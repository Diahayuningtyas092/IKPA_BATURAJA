import streamlit as st
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import io
import numpy as np
import os
import base64
import requests
import re
import calendar
from pathlib import Path
from datetime import datetime
from github import Github
from github import Auth
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

# =========================
# AMBIL PASSWORD DARI SECRETS
# =========================
ADMIN_PASSWORD = st.secrets.get("ADMIN_PASSWORD", "")
if not ADMIN_PASSWORD:
    st.error("ADMIN_PASSWORD belum diset di Streamlit Secrets")
    st.stop()

# Konfigurasi halaman
st.set_page_config(
    page_title="Dashboard IKPA KPPN Baturaja",
    page_icon="📊",
    layout="wide"
)

# define month order map
MONTH_ORDER = {
    'JANUARI': 1,
    'FEBRUARI': 2, 'PEBRUARI': 2, 'PEBRUARY': 2,
    'MARET': 3, 'MAR': 3, 'MRT': 3,
    'APRIL': 4,
    'MEI': 5,
    'JUNI': 6,
    'JULI': 7,
    'AGUSTUS': 8, 'AGUSTUSS': 8,
    'SEPTEMBER': 9, 'SEPT': 9, 'SEP': 9,
    'OKTOBER': 10,
    'NOVEMBER': 11, 'NOPEMBER': 11,
    'DESEMBER': 12
}

# Path ke file template (akan diatur di session state)
TEMPLATE_PATH = r"C:\Users\KEMENKEU\Desktop\INDIKATOR PELAKSANAAN ANGGARAN.xlsx"

# ================================
# INIT SESSION STATE 
# ================================
if "data_storage" not in st.session_state:
    st.session_state.data_storage = {}

if "data_storage_kppn" not in st.session_state:
    st.session_state.data_storage_kppn = {}

if "DATA_DIPA_by_year" not in st.session_state:
    st.session_state.DATA_DIPA_by_year = {}

if "ikpa_dipa_merged" not in st.session_state:
    st.session_state.ikpa_dipa_merged = False

if 'activity_log' not in st.session_state:
    st.session_state.activity_log = [] 

def reset_app_state():
    keys_to_reset = [
        "data_storage",
        "data_storage_kppn",
        "DATA_DIPA_by_year",
        "ikpa_dipa_merged",
        "activity_log",
        "selected_period",
        "main_tab",
        "active_table_tab",
        "period_type",
    ]

    for k in keys_to_reset:
        if k in st.session_state:
            del st.session_state[k]

# -------------------------
# standardize_dipa
# -------------------------
def standardize_dipa(df_raw):

    df = df_raw.copy()
    df.columns = [str(c).strip() for c in df.columns]

    # =============
    # 1) NORMALISASI NAMA KOLOM
    # =============
    def find_col(possible_names):
        for c in df.columns:
            c_norm = re.sub(r'[^A-Z]', '', c.upper())
            for p in possible_names:
                p_norm = re.sub(r'[^A-Z]', '', p.upper())
                if p_norm in c_norm:
                    return c
        return None

    # Cari kolom penting
    col_kode = find_col(["Kode Satker", "Satker"])
    col_nama = find_col(["Nama Satker", "Uraian Satker", "Satker"])
    col_pagu = find_col(["Total Pagu", "Pagu Belanja", "Jumlah"])
    col_tanggal_revisi = find_col(["Tanggal Posting Revisi", "Tanggal Revisi"])
    col_revisi_ke = find_col(["Revisi Terakhir", "Revisi ke"])
    col_no = find_col(["No"])
    col_kementerian = find_col(["Kementerian", "BA", "K/L"])
    col_dipa = find_col(["No Dipa", "Nomor DIPA"])
    col_tanggal_dipa = find_col(["Tanggal Dipa"])
    col_owner = find_col(["Owner"])
    col_stamp = find_col(["Digital Stamp"])
    col_status_history = find_col(["Kode Status History"])
    col_jenis_revisi = find_col(["Jenis Revisi"])

    # =============
    # 2) BUILD OUTPUT
    # =============
    out = pd.DataFrame()

    # KODE SATKER
    if col_kode:
        out["Kode Satker"] = df[col_kode].astype(str).str.extract(r"(\d{6})")[0]
    else:
        out["Kode Satker"] = None

    # NAMA
    if col_nama:
        out["Satker"] = df[col_nama].astype(str).str.replace(r"^\d{6}\s*-?\s*", "", regex=True)
    else:
        out["Satker"] = ""

    # PAGU
    if col_pagu:
        out["Total Pagu"] = (
            df[col_pagu]
            .astype(str)
            .str.replace(r"[^\d\.-]", "", regex=True)
            .astype(float)
            .fillna(0)
            .astype(int)
        )
    else:
        out["Total Pagu"] = 0

    # TANGGAL POSTING REVISI
    if col_tanggal_revisi:
        out["Tanggal Posting Revisi"] = pd.to_datetime(df[col_tanggal_revisi], errors="coerce")
    else:
        out["Tanggal Posting Revisi"] = pd.NaT

    # TAHUN
    out["Tahun"] = out["Tanggal Posting Revisi"].dt.year.fillna(datetime.now().year).astype(int)

    # NO
    if col_no:
        out["NO"] = df[col_no]
    else:
        out["NO"] = range(1, len(df) + 1)

    # KEMENTERIAN
    if col_kementerian:
        out["Kementerian"] = df[col_kementerian].astype(str)
    else:
        if col_dipa:
            out["Kementerian"] = df[col_dipa].astype(str).str.extract(r"DIPA-(\d{3})")[0]
        else:
            out["Kementerian"] = ""

    # REVISI KE
    if col_revisi_ke:
        out["Revisi ke-"] = (
            df[col_revisi_ke]
            .astype(str)
            .str.extract(r"(\d+)")
            .fillna(0)
            .astype(int)
        )
    else:
        out["Revisi ke-"] = 0

    # NO DIPA
    out["No Dipa"] = df[col_dipa].astype(str) if col_dipa else ""

    # TANGGAL DIPA
    out["Tanggal Dipa"] = pd.to_datetime(df[col_tanggal_dipa], errors="coerce") if col_tanggal_dipa else pd.NaT

    # OWNER
    out["Owner"] = df[col_owner].astype(str) if col_owner else ""

    # DIGITAL STAMP
    out["Digital Stamp"] = df[col_stamp].astype(str) if col_stamp else ""

    # Jenis Satker (nanti dihitung di luar)
    out["Jenis Satker"] = ""

    # KODE STATUS HISTORY
    if col_status_history:
        out["Kode Status History"] = df[col_status_history].astype(str)
    else:
        out["Kode Status History"] = ""

    # JENIS REVISI
    if col_jenis_revisi:
        out["Jenis Revisi"] = df[col_jenis_revisi].astype(str)
    else:
        out["Jenis Revisi"] = ""

    # =============
    # 3) FINAL CLEANUP
    # =============
    out = out.dropna(subset=["Kode Satker"])
    out["Kode Satker"] = out["Kode Satker"].astype(str).str.zfill(6)

    # =============
    # 4) SUSUN URUTAN KOLOM
    # =============
    final_order = [
        "Tanggal Posting Revisi",
        "Total Pagu",
        "Jenis Satker",
        "NO",
        "Kementerian",
        "Kode Status History",
        "Jenis Revisi",
        "Revisi ke-",
        "No Dipa",
        "Tanggal Dipa",
        "Owner",
        "Digital Stamp",
    ]

    existing_cols = [c for c in final_order if c in out.columns]
    out = out[existing_cols]

    return out


# Normalize kode satker
def normalize_kode_satker(k, width=6):
    if pd.isna(k):
        return ''
    s = str(k).strip()
    digits = re.findall(r'\d+', s)
    if not digits:
        return ''
    kod = digits[0].zfill(width)
    return kod


st.write("GitHub token loaded:", bool(st.secrets.get("GITHUB_TOKEN")))

def extract_kode_from_satker_field(s, width=6):
    """
    Jika kolom 'Satker' mengandung '001234 – NAMA SATKER', ambil angka di awal.
    Jika hanya angka (sebagai int/str), return padded string.
    """
    if pd.isna(s):
        return ''
    stxt = str(s).strip()
    # cari angka di awal baris (atau angka pertama)
    m = re.match(r'^\s*0*\d+', stxt)
    if m:
        return normalize_kode_satker(m.group(0), width=width)
    # fallback: cari first group of digits anywhere
    m2 = re.search(r'(\d+)', stxt)
    if m2:
        return normalize_kode_satker(m2.group(1), width=width)
    return ''

        
# ===============================
# REGISTER IKPA SATKER (GLOBAL)
# ===============================
def register_ikpa_satker(df_final, month, year, source="Manual"):
    key = (month, str(year))

    df = df_final.copy()

    df["Source"] = source
    df["Period"] = f"{month} {year}"

    MONTH_ORDER = {
        "JANUARI": 1, "FEBRUARI": 2, "MARET": 3, "APRIL": 4,
        "MEI": 5, "JUNI": 6, "JULI": 7, "AGUSTUS": 8,
        "SEPTEMBER": 9, "OKTOBER": 10, "NOVEMBER": 11, "DESEMBER": 12
    }
    df["Period_Sort"] = f"{int(year):04d}-{MONTH_ORDER.get(month, 0):02d}"

    if "Peringkat" not in df.columns:
        df = df.sort_values(
            "Nilai Akhir (Nilai Total/Konversi Bobot)",
            ascending=False
        ).reset_index(drop=True)
        df["Peringkat"] = range(1, len(df) + 1)

    st.session_state.data_storage[key] = df

def find_header_row_by_keyword(uploaded_file, keyword, max_rows=10):
    """
    Mencari baris header Excel berdasarkan keyword kolom
    (contoh: 'Nama KPPN', 'Nama Satker')
    """
    uploaded_file.seek(0)
    preview = pd.read_excel(
        uploaded_file,
        header=None,
        nrows=max_rows
    )

    for i in range(preview.shape[0]):
        row_values = (
            preview.iloc[i]
            .astype(str)
            .str.upper()
            .str.strip()
        )
        if any(keyword.upper() in v for v in row_values):
            return i

    return None

# ===============================
# PARSER IKPA SATKER (INI KUNCI)
# ===============================
def process_excel_file(uploaded_file, upload_year):
    """
    PARSER IKPA SATKER — SATU-SATUNYA YANG BOLEH MEMBACA EXCEL MENTAH
    """
    df_raw = pd.read_excel(uploaded_file, header=None)

    # Ambil bulan dari baris atas
    month_text = str(df_raw.iloc[1, 0])
    month = month_text.split(":")[-1].strip().upper()

    # Data mulai baris ke-5
    df_data = df_raw.iloc[4:].reset_index(drop=True)
    df_data.columns = range(len(df_data.columns))

    processed_rows = []
    i = 0
    while i + 3 < len(df_data):

        nilai = df_data.iloc[i]
        bobot = df_data.iloc[i + 1]
        nilai_akhir = df_data.iloc[i + 2]
        nilai_aspek = df_data.iloc[i + 3]

        row = {
            "No": nilai[0],
            "Kode KPPN": str(nilai[1]).strip("'"),
            "Kode BA": str(nilai[2]).strip("'"),
            "Kode Satker": str(nilai[3]).strip("'"),
            "Uraian Satker": nilai[4],

            "Kualitas Perencanaan Anggaran": nilai_aspek[6],
            "Kualitas Pelaksanaan Anggaran": nilai_aspek[8],
            "Kualitas Hasil Pelaksanaan Anggaran": nilai_aspek[12],

            "Revisi DIPA": nilai[6],
            "Deviasi Halaman III DIPA": nilai[7],
            "Penyerapan Anggaran": nilai[8],
            "Belanja Kontraktual": nilai[9],
            "Penyelesaian Tagihan": nilai[10],
            "Pengelolaan UP dan TUP": nilai[11],
            "Capaian Output": nilai[12],

            "Nilai Total": nilai[13],
            "Konversi Bobot": nilai[14],
            "Dispensasi SPM (Pengurang)": nilai[15],
            "Nilai Akhir (Nilai Total/Konversi Bobot)": nilai[16],

            "Bulan": month,
            "Tahun": upload_year
        }

        processed_rows.append(row)
        i += 4

    df_final = pd.DataFrame(processed_rows)
    return df_final, month, upload_year

# ===============================
# REPROCESS ALL IKPA SATKER
# ===============================
def reprocess_all_ikpa_satker():
    with st.spinner("🔄 Memproses ulang seluruh IKPA Satker..."):
        load_data_from_github()
        st.session_state.ikpa_dipa_merged = False


def process_excel_file_kppn(uploaded_file, year):
    try:
        import pandas as pd

        # ===============================
        # HELPER AMAN AKSES INDEX
        # ===============================
        def safe(row, idx):
            return row[idx] if idx < len(row) and pd.notna(row[idx]) else 0

        # ===============================
        # BACA FILE
        # ===============================
        df_raw = pd.read_excel(uploaded_file, header=None)

        # ===============================
        # DETEKSI BULAN
        # ===============================
        month = "UNKNOWN"
        if df_raw.shape[0] > 1:
            text = str(df_raw.iloc[1, 0]).upper()
            MONTH_MAP = {
                "JAN": "JANUARI", "JANUARI": "JANUARI",
                "FEB": "FEBRUARI", "FEBRUARI": "FEBRUARI",
                "MAR": "MARET", "MARET": "MARET",
                "APR": "APRIL", "APRIL": "APRIL",
                "MEI": "MEI",
                "JUN": "JUNI", "JUNI": "JUNI",
                "JUL": "JULI", "JULI": "JULI",
                "AGT": "AGUSTUS", "AGS": "AGUSTUS", "AGUSTUS": "AGUSTUS",
                "SEP": "SEPTEMBER", "SEPTEMBER": "SEPTEMBER",
                "OKT": "OKTOBER", "OKTOBER": "OKTOBER",
                "NOV": "NOVEMBER", "NOVEMBER": "NOVEMBER",
                "DES": "DESEMBER", "DESEMBER": "DESEMBER"
            }
            for k, v in MONTH_MAP.items():
                if k in text:
                    month = v
                    break

        # ===============================
        # DATA MULAI BARIS KE-5
        # ===============================
        df_data = df_raw.iloc[4:].reset_index(drop=True)
        df_data.columns = range(len(df_data.columns))

        processed_rows = []
        i = 0

        while i + 3 < len(df_data):
            nilai_row = df_data.iloc[i]
            nilai_aspek_row = df_data.iloc[i + 3]

            row_data = {
                # IDENTITAS
                "No": safe(nilai_row, 0),
                "Kode KPPN": str(safe(nilai_row, 1)).replace("'", "").strip(),
                "Nama KPPN": safe(nilai_row, 2),

                # KUALITAS (DARI NILAI ASPEK)
                "Kualitas Perencanaan Anggaran": safe(nilai_aspek_row, 4),
                "Kualitas Pelaksanaan Anggaran": safe(nilai_aspek_row, 6),
                "Kualitas Hasil Pelaksanaan Anggaran": safe(nilai_aspek_row, 10),

                # INDIKATOR
                "Revisi DIPA": safe(nilai_row, 4),
                "Deviasi Halaman III DIPA": safe(nilai_row, 5),
                "Penyerapan Anggaran": safe(nilai_row, 6),
                "Belanja Kontraktual": safe(nilai_row, 7),
                "Penyelesaian Tagihan": safe(nilai_row, 8),
                "Pengelolaan UP dan TUP": safe(nilai_row, 9),
                "Capaian Output": safe(nilai_row, 10),

                # NILAI AKHIR
                "Nilai Total": safe(nilai_row, 11),
                "Konversi Bobot": safe(nilai_row, 12),
                "Dispensasi SPM (Pengurang)": safe(nilai_row, 13),
                "Nilai Akhir (Nilai Total/Konversi Bobot)": safe(nilai_row, 14),

                # METADATA
                "Bulan": month,
                "Tahun": year,
                "Source": "Upload"
            }

            processed_rows.append(row_data)
            i += 4  # 🔑 POLA IKPA (Nilai, Bobot, Nilai Akhir, Nilai Aspek)

        df = pd.DataFrame(processed_rows)

        # ===============================
        # NUMERIC CAST
        # ===============================
        numeric_cols = [
            "Kualitas Perencanaan Anggaran",
            "Kualitas Pelaksanaan Anggaran",
            "Kualitas Hasil Pelaksanaan Anggaran",
            "Revisi DIPA",
            "Deviasi Halaman III DIPA",
            "Penyerapan Anggaran",
            "Belanja Kontraktual",
            "Penyelesaian Tagihan",
            "Pengelolaan UP dan TUP",
            "Capaian Output",
            "Nilai Total",
            "Konversi Bobot",
            "Dispensasi SPM (Pengurang)",
            "Nilai Akhir (Nilai Total/Konversi Bobot)"
        ]

        for col in numeric_cols:
            if col in df.columns:
                df[col] = pd.to_numeric(df[col], errors="coerce")

        # ===============================
        # RANKING
        # ===============================
        df = df.sort_values(
            "Nilai Akhir (Nilai Total/Konversi Bobot)",
            ascending=False
        ).reset_index(drop=True)

        df["Peringkat"] = df.index + 1

        return df, month, year

    except Exception as e:
        st.error(f"❌ Error memproses IKPA KPPN: {e}")
        return None, None, None


# ============================================================
# PARSER DIPA 
# ============================================================
#Parser Perbaikan DIPA
def parse_dipa(df_raw):
    import pandas as pd
    import re
    from datetime import datetime

    # ====== 1. Hapus baris kosong ======
    df = df_raw.dropna(how="all").reset_index(drop=True)

    # ====== 2. Cari baris header yang BENAR ======
    header_row = None
    for i in range(min(10, len(df))):
        row_str = " ".join(df.iloc[i].astype(str).str.upper().tolist())
        if (
            "NO" in row_str
            and "SATKER" in row_str
            and "DIPA" in row_str
        ):
            header_row = i
            break

    # Jika tidak ketemu → fallback (baris 2 biasanya)
    if header_row is None:
        header_row = 2

    # ====== 3. Set header ======
    df.columns = df.iloc[header_row]
    df = df.iloc[header_row + 1:].reset_index(drop=True)

    # ====== 4. Normalisasi kolom ======
    def get(names):
        for c in df.columns:
            cc = str(c).upper().replace(".", "").strip()
            for n in names:
                if n in cc:
                    return c
        return None

    col_no     = get(["NO"])
    col_satker = get(["SATKER"])
    col_nama   = get(["NAMA SATKER"])
    col_dipa   = get(["NO DIPA"])
    col_pagu   = get(["PAGU", "TOTAL PAGU"])
    col_tgl    = get(["TANGGAL POSTING", "TANGGAL REVISI"])
    col_rev    = get(["REVISI TERAKHIR", "REVISI KE"])
    col_tgl_dipa = get(["TANGGAL DIPA"])
    col_owner  = get(["OWNER"])
    col_stamp  = get(["STAMP"])
    col_status = get(["STATUS", "HISTORY"])

    out = pd.DataFrame()

    # NO
    out["NO"] = df[col_no] if col_no else range(1, len(df)+1)

    # Kode Satker
    out["Kode Satker"] = df[col_satker].astype(str).str.extract(r"(\d{6})")[0]

    # Nama Satker
    if col_nama:
        out["Satker"] = df[col_nama].astype(str)
    else:
        out["Satker"] = df[col_satker].astype(str).str.replace(r"^\d{6}\s*-?\s*", "", regex=True)

    # Total Pagu
    out["Total Pagu"] = (
        df[col_pagu].astype(str)
        .str.replace(r"[^\d\-\.]", "", regex=True)
        .replace("", "0")
        .astype(float)
    ) if col_pagu else 0

    # No DIPA
    out["No Dipa"] = df[col_dipa].astype(str)

    # Kementerian (BA)
    out["Kementerian"] = out["No Dipa"].str.extract(r"DIPA-(\d{3})")[0].fillna("")

    # Kode Status History -> BXX
    out["Kode Status History"] = (
        "B" + out["No Dipa"].str.extract(r"DIPA-\d{3}\.(\d{2})")[0].fillna("00")
    )

    # Revisi ke
    if col_rev:
        r = df[col_rev].astype(str).str.extract(r"(\d+)")[0].fillna(0).astype(int)
        out["Revisi ke-"] = r
        out["Jenis Revisi"] = r.apply(lambda x: "DIPA_REVISI" if x > 0 else "ANGKA_DASAR")
    else:
        out["Revisi ke-"] = 0
        out["Jenis Revisi"] = "ANGKA_DASAR"

    # Tanggal DIPA
    out["Tanggal Dipa"] = (
        pd.to_datetime(df[col_tgl_dipa], errors="coerce")
        if col_tgl_dipa else pd.NaT
    )

    # Tanggal Posting Revisi -> format dd-mm-yyyy
    out["Tanggal Posting Revisi"] = (
        pd.to_datetime(df[col_tgl], format="%d-%m-%Y", errors="coerce")
        if col_tgl else pd.NaT
    )

    # Tahun
    out["Tahun"] = (
        out["Tanggal Posting Revisi"].dt.year
            .fillna(datetime.now().year)
            .astype(int)
    )

    # Owner (default untuk 2022–2024)
    out["Owner"] = (
        df[col_owner].astype(str)
        if col_owner else "UNIT"
    )

    # Digital Stamp (default untuk 2022–2024)
    out["Digital Stamp"] = (
        df[col_stamp].astype(str)
        if col_stamp else "0000000000000000"
    )

    # Jenis Satker (auto berdasarkan pagu)
    out["Jenis Satker"] = out["Total Pagu"].apply(
        lambda x: "Satker Besar" if x >= 10_000_000_000 
        else ("Satker Sedang" if x >= 1_000_000_000 else "Satker Kecil")
    )

    out = out.dropna(subset=["Kode Satker"])
    out["Kode Satker"] = out["Kode Satker"].astype(str).str.zfill(6)

    return out


# ============================================================
# FUNGSI HELPER: Load Data DIPA dari GitHub
# ============================================================
def load_DATA_DIPA_from_github():
    token = st.secrets.get("GITHUB_TOKEN")
    repo_name = st.secrets.get("GITHUB_REPO")

    if not token or not repo_name:
        st.error("❌ GitHub token / repo tidak ditemukan.")
        return False

    try:
        g = Github(auth=Auth.Token(token))
        repo = g.get_repo(repo_name)
    except:
        st.error("❌ Gagal koneksi GitHub.")
        return False

    try:
        files = repo.get_contents("DATA_DIPA")
    except:
        st.error("❌ Folder DATA_DIPA tidak ditemukan di GitHub.")
        return False

    pattern = re.compile(r"^DIPA[_-]?(\d{4})\.xlsx$", re.IGNORECASE)

    st.session_state.DATA_DIPA_by_year = {}
    loaded_years = []

    for f in files:
        match = pattern.match(f.name)
        if not match:
            continue

        tahun = int(match.group(1))

        try:
            raw = base64.b64decode(f.content)
            df_raw = pd.read_excel(io.BytesIO(raw), header=None)

            # GUNAKAN PARSER BARU
            df_parsed = parse_dipa(df_raw)

            # Set tahun
            df_parsed["Tahun"] = tahun

            # Simpan
            st.session_state.DATA_DIPA_by_year[tahun] = df_parsed
            loaded_years.append(str(tahun))

        except Exception as e:
            st.warning(f"⚠️ DIPA {tahun} gagal diproses: {e}")

    if loaded_years:
        st.success("✅ DIPA berhasil dimuat: " + ", ".join(loaded_years))
    else:
        st.error("❌ Tidak ada data DIPA yang dapat diproses.")

    return True


# Save any file (Excel/template) to your GitHub repo
def save_file_to_github(content_bytes, filename, folder):
    token = st.secrets["GITHUB_TOKEN"]
    repo_name = st.secrets["GITHUB_REPO"]

    g = Github(auth=Auth.Token(token))
    repo = g.get_repo(repo_name)
    

    # 1️⃣ buat path full
    path = f"{folder}/{filename}"

    try:
        # 2️⃣ cek apakah file sudah ada
        existing = repo.get_contents(path)
        repo.update_file(existing.path, f"Update {filename}", content_bytes, existing.sha)
    except Exception:
        # 3️⃣ jika folder tidak ada → buat file pertama
        repo.create_file(path, f"Create {filename}", content_bytes)
        

# ============================
#  LOAD DATA IKPA DARI GITHUB
# ============================
def load_data_from_github():
    """
    Load IKPA Satker dari GitHub (/data).
    HANYA file hasil proses (df_final) yang diterima.
    Data manual TIDAK akan ditimpa.
    """

    token = st.secrets.get("GITHUB_TOKEN")
    repo_name = st.secrets.get("GITHUB_REPO")

    if not token or not repo_name:
        st.error("❌ Gagal mengakses GitHub: token/repo tidak ditemukan.")
        st.stop()
        return

    g = Github(auth=Auth.Token(token))
    repo = g.get_repo(repo_name)

    try:
        contents = repo.get_contents("data")
    except Exception:
        st.warning("📁 Folder 'data' belum ada di GitHub.")
        return

    # JANGAN RESET data_storage
    if "data_storage" not in st.session_state:
        st.session_state.data_storage = {}

    REQUIRED_COLUMNS = [
        "No", "Kode KPPN", "Kode BA", "Kode Satker", "Uraian Satker",
        "Kualitas Perencanaan Anggaran",
        "Kualitas Pelaksanaan Anggaran",
        "Kualitas Hasil Pelaksanaan Anggaran",
        "Revisi DIPA", "Deviasi Halaman III DIPA",
        "Penyerapan Anggaran", "Belanja Kontraktual",
        "Penyelesaian Tagihan", "Pengelolaan UP dan TUP",
        "Capaian Output",
        "Nilai Total", "Konversi Bobot",
        "Dispensasi SPM (Pengurang)",
        "Nilai Akhir (Nilai Total/Konversi Bobot)",
        "Bulan", "Tahun"
    ]

    MONTH_ORDER = {
        "JANUARI": 1, "FEBRUARI": 2, "MARET": 3, "APRIL": 4,
        "MEI": 5, "JUNI": 6, "JULI": 7, "AGUSTUS": 8,
        "SEPTEMBER": 9, "OKTOBER": 10, "NOVEMBER": 11, "DESEMBER": 12
    }

    loaded_count = 0

    for file in contents:
        if not file.name.endswith(".xlsx"):
            continue

        try:
            decoded = base64.b64decode(file.content)
            df = pd.read_excel(io.BytesIO(decoded))

            # VALIDASI HASIL PROSES
            if not all(col in df.columns for col in REQUIRED_COLUMNS):
                continue

            month = str(df["Bulan"].iloc[0]).upper()
            year = str(df["Tahun"].iloc[0])
            key = (month, year)

            # ❗ JIKA SUDAH ADA (MANUAL), LEWATI
            if key in st.session_state.data_storage:
                continue

            # NORMALISASI
            df["Bulan"] = month
            df["Tahun"] = year

            if "Kode Satker" in df.columns:
                df["Kode Satker"] = (
                    df["Kode Satker"]
                    .astype(str)
                    .apply(normalize_kode_satker)
                )

            try:
                df = apply_reference_short_names(df)
            except:
                pass

            try:
                df = create_satker_column(df)
            except:
                pass

            numeric_cols = [
                "Nilai Akhir (Nilai Total/Konversi Bobot)",
                "Nilai Total", "Konversi Bobot",
                "Revisi DIPA", "Deviasi Halaman III DIPA",
                "Penyerapan Anggaran", "Belanja Kontraktual",
                "Penyelesaian Tagihan", "Pengelolaan UP dan TUP",
                "Capaian Output",
                "Kualitas Perencanaan Anggaran",
                "Kualitas Pelaksanaan Anggaran",
                "Kualitas Hasil Pelaksanaan Anggaran",
            ]

            for col in numeric_cols:
                if col in df.columns:
                    df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)

            month_num = MONTH_ORDER.get(month, 0)

            df["Source"] = "GitHub"
            df["Period"] = f"{month} {year}"
            df["Period_Sort"] = f"{int(year):04d}-{month_num:02d}"

            if "Peringkat" not in df.columns:
                df = df.sort_values(
                    "Nilai Akhir (Nilai Total/Konversi Bobot)",
                    ascending=False
                ).reset_index(drop=True)
                df["Peringkat"] = range(1, len(df) + 1)

            st.session_state.data_storage[key] = df
            loaded_count += 1

        except Exception as e:
            st.error(f"❌ Gagal memuat {file.name}: {e}")

    st.success(f"✅ {loaded_count} file IKPA Satker dimuat dari GitHub.")


# ============================
#  BACA TEMPLATE FILE
# ============================
def get_template_file():
    try:
        if Path(TEMPLATE_PATH).exists():
            with open(TEMPLATE_PATH, "rb") as f:
                return f.read()
        else:
            if "template_file" in st.session_state:
                return st.session_state.template_file
            return None
    except Exception as e:
        st.error(f"Error membaca template: {e}")
        return None

# Fungsi visualisasi podium/bintang
def create_ranking_chart(df, title, top=True, limit=10):
    """
    Membuat visualisasi ranking dengan bar chart horizontal yang menarik
    (Sekarang menggunakan kolom 'Satker' untuk label agar unik)
    """
    if top:
        df_sorted = df.nlargest(limit, 'Nilai Akhir (Nilai Total/Konversi Bobot)')
        color_scale = 'Greens'
        emoji = '🏆'
    else:
        df_sorted = df.nsmallest(limit, 'Nilai Akhir (Nilai Total/Konversi Bobot)')
        color_scale = 'Reds'
        emoji = '⚠️'
    
    fig = go.Figure()
    
    colors = px.colors.sequential.Greens if top else px.colors.sequential.Reds
    
    # use 'Satker' for y labels to keep them unique
    fig.add_trace(go.Bar(
    y=df_filtered['Satker'],
    x=df_filtered[column],
    orientation='h',
    marker=dict(
        color=df_filtered[column],
        colorscale='OrRd_r',
        showscale=True,
        cmin=min_val,
        cmax=max_val,
    ),
    text=df_filtered[column].round(2),
    textposition='outside',
    hovertemplate='<b>%{y}</b><br>Nilai: %{x:.2f}<extra></extra>'
))
    
    fig.update_layout(
        title=f"{emoji} {title}",
        xaxis_title="Nilai Akhir",
        yaxis_title="",
        height=max(400, limit * 40),
        yaxis={'categoryorder': 'total ascending' if not top else 'total descending'},
        showlegend=False
    )
    # ============================
    # Rotated labels 45° di bawah
    # ============================
    annotations = []
    y_positions = list(range(len(df_filtered)))

    for i, satker in enumerate(df_filtered['Satker']):
        annotations.append(dict(
        x=df_filtered[column].min() - 3,
        y=i,
        text=satker,
        xanchor="right",
        yanchor="middle",
        showarrow=False,
        textangle=45,
        font=dict(size=10),
    ))

    fig.update_layout(annotations=annotations)

    # Sembunyikan label Y-axis
    fig.update_yaxes(showticklabels=False)

    return fig

# ============================================================
# Improved Problem Chart (with sorting, sliders, and filters)
# ============================================================
def get_top_bottom(df, n=10, top=True):
    if df.empty:
        return df
    return (
        df.nlargest(n, 'Nilai Akhir (Nilai Total/Konversi Bobot)')
        if top else
        df.nsmallest(n, 'Nilai Akhir (Nilai Total/Konversi Bobot)')
    )


def make_column_chart(data, title, color_scale, y_min, y_max):
    if data.empty:
        return None

    fig = px.bar(
        data.sort_values("Nilai Akhir (Nilai Total/Konversi Bobot)"),
        x="Nilai Akhir (Nilai Total/Konversi Bobot)",
        y="Satker",
        orientation="h",
        color="Nilai Akhir (Nilai Total/Konversi Bobot)",
        color_continuous_scale=color_scale,
        title=title
    )

    fig.update_layout(
        xaxis_range=[y_min, y_max],
        xaxis_title="Nilai IKPA",
        yaxis_title="",
        height=450,
        margin=dict(l=10, r=10, t=40, b=20),
        coloraxis_showscale=False,
        showlegend=False
    )

    fig.update_traces(
        texttemplate="%{x:.2f}",
        textposition="outside",
        hovertemplate="<b>%{y}</b><br>Nilai: %{x:.2f}<extra></extra>"
    )

    return fig

def safe_chart(df, title, top=True, color="Greens", y_min=0, y_max=110):
    if df is None or df.empty:
        st.info("Tidak ada data.")
        return

    chart_df = get_top_bottom(df, 10, top)
    if chart_df is None or chart_df.empty:
        st.info("Tidak ada data.")
        return

    fig = make_column_chart(chart_df, "", color, y_min, y_max)
    if fig:
        st.plotly_chart(fig, use_container_width=True)

# ============================================================
# Problem Chart untuk Dashboard Internal
# ============================================================
def create_problem_chart(df, column, threshold, title, comparison='less', y_min=None, y_max=None, show_yaxis=True):

    if comparison == 'less':
        df_filtered = df[df[column] < threshold]
    elif comparison == 'greater':
        df_filtered = df[df[column] > threshold]
    else:
        df_filtered = df.copy()

    # Jika hasil filter kosong → Cegah error
    if df_filtered.empty:
        df_filtered = df.head(1)

    df_filtered = df_filtered.sort_values(by=column, ascending=False)

    # Ambil nilai range untuk colormap
    min_val = df_filtered[column].min()
    max_val = df_filtered[column].max()

    fig = go.Figure()
    fig.add_trace(go.Bar(
        x=df_filtered['Satker'],
        y=df_filtered[column],
        marker=dict(
            color=df_filtered[column],
            colorscale='OrRd_r',
            showscale=True,
            cmin=min_val,
            cmax=max_val,
        ),
        text=df_filtered[column].round(2),
        textposition='outside',
        textangle=0,
        textfont=dict(family="Arial Black", size=12), 
        hovertemplate='<b>%{x}</b><br>Nilai: %{y:.2f}<extra></extra>'
    ))

    # Garis target threshold (tidak berubah)
    fig.add_hline(
        y=threshold,
        line_dash="dash",
        line_color="red",
        annotation_text=f"Target: {threshold}",
        annotation_position="top right"
    )

    # Bold judul dan label axis
    fig.update_layout(
        xaxis=dict(
        tickangle=-45,
        tickmode='linear',
        tickfont=dict(family="Arial Black", size=10),
        automargin=True
    ),
    yaxis=dict(
        tickfont=dict(family="Arial Black", size=11)
        ),
        height=600,
        margin=dict(l=50, r=20, t=80, b=200),
        showlegend=False,
    )

    if not show_yaxis:
        fig.update_yaxes(showticklabels=False)

    return fig
# ===============================================
# Helper to apply reference short names (Simplified)
# ===============================================
def apply_reference_short_names(df):
    """
    Simple version: apply reference short names to dataframe.
    - Adds 'Uraian Satker-RINGKAS' (from reference 'Uraian Satker-SINGKAT' when available,
      otherwise falls back to original 'Uraian Satker').
    - Performs basic normalization on 'Kode Satker' before merging.
    - Minimal user messages (no Excel/CSV creation, no verbose debugging).
    """
    # Defensive copy
    df = df.copy()

    # Ensure period columns exist
    if 'Bulan' not in df.columns:
        df['Bulan'] = ''
    if 'Tahun' not in df.columns:
        df['Tahun'] = ''

    # If no reference in session, fallback silently to original names
    if 'reference_df' not in st.session_state or st.session_state.reference_df is None:
        if 'Uraian Satker-RINGKAS' not in df.columns:
            df['Uraian Satker-RINGKAS'] = df.get('Uraian Satker', '')
        # also keep a final fallback column for compatibility
        df['Uraian Satker Final'] = df.get('Uraian Satker', '')
        return df

    # Copy reference
    ref = st.session_state.reference_df.copy()

    # Normalize Kode Satker if column exists; else create empty codes to avoid crashes
    if 'Kode Satker' in df.columns:
        df['Kode Satker'] = df['Kode Satker'].apply(normalize_kode_satker)
    else:
        df['Kode Satker'] = ''

    if 'Kode Satker' in ref.columns:
        ref['Kode Satker'] = ref['Kode Satker'].apply(normalize_kode_satker)
    else:
        # If reference has no Kode Satker, cannot match — fallback
        if 'Uraian Satker-RINGKAS' not in df.columns:
            df['Uraian Satker-RINGKAS'] = df.get('Uraian Satker', '')
        df['Uraian Satker Final'] = df.get('Uraian Satker', '')
        return df

    # Ensure kode fields are strings and stripped
    df['Kode Satker'] = df['Kode Satker'].astype(str).str.strip()
    ref['Kode Satker'] = ref['Kode Satker'].astype(str).str.strip()

    # If the reference does not contain the expected short-name column, fallback
    if 'Uraian Satker-SINGKAT' not in ref.columns:
        if 'Uraian Satker-RINGKAS' not in df.columns:
            df['Uraian Satker-RINGKAS'] = df.get('Uraian Satker', '')
        df['Uraian Satker Final'] = df.get('Uraian Satker', '')
        return df

    # Perform the merge and create final short-name column; keep it simple and robust
    try:
        df_merged = df.merge(
            ref[['Kode Satker', 'Uraian Satker-SINGKAT']].rename(columns={'Uraian Satker-SINGKAT': 'Uraian Satker-RINGKAS'}),
            on='Kode Satker',
            how='left',
            indicator=False
        )

        # Create final name column using reference when available, otherwise fallback to original
        df_merged['Uraian Satker-RINGKAS'] = df_merged['Uraian Satker-RINGKAS'].fillna(
            df_merged.get('Uraian Satker', '')
        )

        # Keep a generic final field for backward compatibility
        df_merged['Uraian Satker Final'] = df_merged['Uraian Satker-RINGKAS']

        # Drop the reference short-name column in case it remains under other names
        df_merged = df_merged.drop(columns=['Uraian Satker-SINGKAT'], errors='ignore')

        return df_merged

    except Exception as e:
        # Minimal error notification and fallback
        st.error(f"❌ Gagal menerapkan nama singkat untuk periode {df.get('Bulan', [''])[0]} {df.get('Tahun', [''])[0]}: {e}")
        if 'Uraian Satker-RINGKAS' not in df.columns:
            df['Uraian Satker-RINGKAS'] = df.get('Uraian Satker', '')
        df['Uraian Satker Final'] = df.get('Uraian Satker', '')
        return df

# ===============================================
# UPDATED: Helper function to create Satker column consistently
# ===============================================
def create_satker_column(df):
    """
    Creates 'Satker' column consistently across all data sources.
    Should be called after apply_reference_short_names().
    """
    if 'Uraian Satker-RINGKAS' not in df.columns:
        # fallback to older field names
        if 'Uraian Satker Final' in df.columns:
            df['Uraian Satker-RINGKAS'] = df['Uraian Satker Final']
        else:
            df['Uraian Satker-RINGKAS'] = df.get('Uraian Satker', '')

    # Create Satker display using ringkas
    df['Satker'] = (
        df['Uraian Satker-RINGKAS'].astype(str) + 
        ' (' + df['Kode Satker'].astype(str) + ')'
    )
    # Keep backward compatible column
    df['Uraian Satker Final'] = df['Uraian Satker-RINGKAS']
    return df

# BAGIAN 4 CHART DASHBOARD UTAMA
def safe_chart(
    df_part,
    jenis,
    top=True,
    color="Greens",
    y_min=None,
    y_max=None,
    thin_bar=False
):
    if df_part is None or df_part.empty:
        st.info("Tidak ada data.")
        return

    df = df_part.copy()

    # pastikan kolom Satker
    if "Satker" not in df.columns:
        if "Uraian Satker-RINGKAS" in df.columns and "Kode Satker" in df.columns:
            df["Satker"] = (
                df["Uraian Satker-RINGKAS"].astype(str)
                + " (" + df["Kode Satker"].astype(str) + ")"
            )
        else:
            st.warning("Kolom Satker tidak tersedia.")
            return

    # deteksi kolom nilai
    kandidat_ikpa = [
        "Nilai Akhir (Nilai Total/Konversi Bobot)",
        "Nilai Total/Konversi Bobot",
        "Nilai Total"
    ]

    nilai_col = next((c for c in kandidat_ikpa if c in df.columns), None)
    if nilai_col is None:
        st.warning("Kolom nilai IKPA tidak ditemukan.")
        return

    df[nilai_col] = pd.to_numeric(df[nilai_col], errors="coerce")
    df = df.dropna(subset=[nilai_col])

    if df.empty:
        st.info("Tidak ada data valid.")
        return

    df_sorted = (
        df.sort_values(nilai_col, ascending=not top)
          .head(10)
          .sort_values(nilai_col, ascending=True)
    )

    fig = px.bar(
        df_sorted,
        x=nilai_col,
        y="Satker",
        orientation="h",
        color=nilai_col,
        color_continuous_scale=color,
        text=nilai_col
    )

    fig.update_traces(
        texttemplate="%{text:.2f}",
        textposition="outside",
        cliponaxis=False,
        width=0.6 if thin_bar else 0.8
    )

    fig.update_layout(
        height=max(250, len(df_sorted) * 30),
        margin=dict(l=10, r=10, t=10, b=10),
        xaxis=dict(range=[y_min, y_max] if y_min and y_max else None),
        showlegend=False,
        coloraxis_showscale=False
    )

    st.plotly_chart(fig, use_container_width=True)


# HALAMAN 1: DASHBOARD UTAMA (REVISED)
def page_dashboard():
    st.title("📊 Dashboard Utama IKPA Satker Mitra KPPN Baturaja")
    
    st.markdown("""
    <style>
    /* Warna tombol popover */
    div[data-testid="stPopover"] button {
        background-color: #FFF9E6 !important;
        border: 1px solid #E6C200 !important;
        color: #664400 !important;
    }
    div[data-testid="stPopover"] button:hover {
        background-color: #FFE4B5 !important;
        color: black !important;
    }
    button[data-testid="baseButton"][kind="popover"] {
        background-color: #FFF9E6 !important;
        border: 1px solid #E6C200 !important;
        color: #664400 !important;
    }
    button[data-testid="baseButton"][kind="popover"]:hover {
        background-color: #FFE4B5 !important;
    }
    </style>
    """, unsafe_allow_html=True)

    # protect against missing data_storage
    if not st.session_state.get('data_storage'):
        st.warning("⚠️ Belum ada data yang diunggah. Silakan unggah data melalui halaman Admin.")
        return

    # Dapatkan data terbaru
    try:
        all_periods = sorted(
            st.session_state.data_storage.keys(),
            key=lambda x: (int(x[1]), MONTH_ORDER.get(x[0].upper(), 0)),
            reverse=True
        )
    except Exception:
        st.warning("⚠️ Format periode pada data tidak sesuai. Periksa struktur data di session_state.data_storage.")
        return

    if not all_periods:
        st.warning("⚠️ Belum ada data yang tersedia.")
        return

    # ---------------------------
    # Ensure a selected_period and df exist BEFORE any branch uses df
    # ---------------------------
    if "selected_period" not in st.session_state:
        st.session_state.selected_period = all_periods[0]

    # safe fetch of df for the selected_period (always a tuple key like ('JANUARI','2025'))
    selected_period_key = st.session_state.get("selected_period", all_periods[0])
    df = st.session_state.data_storage.get(selected_period_key, None)

    if df is None:
        st.warning(f"⚠️ Data untuk periode {selected_period_key} tidak ditemukan. Periksa st.session_state.data_storage keys.")
        # show available keys to help debugging (optional - remove if sensitive)
        st.write("Periode yang tersedia:", list(st.session_state.data_storage.keys()))
        return

    # ensure main_tab state exists
    if "main_tab" not in st.session_state:
        st.session_state.main_tab = "🎯 Highlights"

    # ---------- persistent main tab ----------
    main_tab = st.radio(
        "Pilih Bagian Dashboard",
        ["🎯 Highlights", "📋 Data Detail Satker"],
        key="main_tab_choice",
        horizontal=True
    )
    st.session_state["main_tab"] = main_tab

    # -------------------------
    # HIGHLIGHTS
    # -------------------------
    if main_tab == "🎯 Highlights":
        st.markdown("## 🎯 Highlights Kinerja Satker")

        # -------------------------
        # Pilih Periode
        # -------------------------
        selected_period = st.selectbox(
            "Pilih Periode",
            options=all_periods,
            index=0,
            format_func=lambda x: f"{x[0].capitalize()} {x[1]}",
            key="select_period_main"
        )
        df = st.session_state.data_storage.get(selected_period)

        # ===============================
        # Validasi DF
        # ===============================
        if df is None or df.empty:
            st.warning("Data IKPA belum tersedia.")
            st.stop()

        # ===============================
        # Pastikan kolom Jenis Satker ada
        # ===============================
        if 'Jenis Satker' not in df.columns:
            df['Jenis Satker'] = 'TIDAK TERKLASIFIKASI'
        else:
            df['Jenis Satker'] = df['Jenis Satker'].fillna('TIDAK TERKLASIFIKASI')

        # ===============================
        # NORMALISASI JENIS SATKER
        # ===============================
        df['Jenis Satker'] = (
            df['Jenis Satker']
            .str.upper()
            .str.replace('SATKER ', '', regex=False)
            .str.strip()
        )

        # ===============================
        # Filter Satker
        # ===============================
        VALID_JENIS = ['KECIL', 'SEDANG', 'BESAR']
        df = df[df['Jenis Satker'].isin(VALID_JENIS)]
        df_kecil  = df[df['Jenis Satker'] == 'KECIL']
        df_sedang = df[df['Jenis Satker'] == 'SEDANG']
        df_besar  = df[df['Jenis Satker'] == 'BESAR']

        # ===============================
        # METRIK UTAMA
        # ===============================
        nilai_col = 'Nilai Akhir (Nilai Total/Konversi Bobot)'

        avg_score = df[nilai_col].mean()
        perfect_df = df[df[nilai_col] == 100]
        below89_df = df[df[nilai_col] < 89]

        # Pastikan kolom Satker tersedia
        def make_satker_col(dd):
            if 'Satker' in dd.columns:
                return dd
            uraian = dd.get('Uraian Satker-RINGKAS', dd.index.astype(str))
            kode = dd.get('Kode Satker', '')
            dd = dd.copy()
            dd['Satker'] = uraian.astype(str) + " (" + kode.astype(str) + ")"
            return dd

        perfect_df = make_satker_col(perfect_df)
        below89_df = make_satker_col(below89_df)

        jumlah_100 = len(perfect_df)
        jumlah_below = len(below89_df)

        # Tampilan metrik
        col1, col2, col3, col4 = st.columns(4)
        with col1:
            st.metric("📋 Total Satker", len(df))
        with col2:
            st.metric("📈 Rata-rata Nilai", f"{avg_score:.2f}")
        with col3:
            st.metric("⭐ Nilai 100", jumlah_100)
            with st.popover("Lihat daftar satker"):
                if jumlah_100 == 0:
                    st.write("Tidak ada satker dengan nilai 100.")
                else:
                    display_df = perfect_df[['Satker']].reset_index(drop=True)
                    display_df.insert(0, 'No', range(1, len(display_df) + 1))
                    st.dataframe(
                        display_df,
                        use_container_width=True,
                        hide_index=True,
                        height=min(400, len(display_df) * 35 + 38)
                    )
        with col4:
            st.metric("⚠️ Nilai < 89 (Predikat Belum Baik)", jumlah_below)
            with st.popover("Lihat daftar satker"):
                if jumlah_below == 0:
                    st.write("Tidak ada satker dengan nilai < 89.")
                else:
                    display_df = below89_df[['Satker']].reset_index(drop=True)
                    display_df.insert(0, 'No', range(1, len(display_df) + 1))
                    st.dataframe(
                        display_df,
                        use_container_width=True,
                        hide_index=True,
                        height=min(400, len(display_df) * 35 + 38)
                    )

        # ===============================
        # Kontrol Skala Chart
        # ===============================
        st.markdown("###### Atur Skala Nilai (Sumbu Y)")
        col_min, col_max = st.columns(2)
        with col_min:
            y_min = st.slider("Nilai Minimum (Y-Axis)", 0, 50, 50, 1, key="high_ymin")
        with col_max:
            y_max = st.slider("Nilai Maksimum (Y-Axis)", 51, 110, 110, 1, key="high_ymax")

        # ===============================
        # CHART 6 MUAT DALAM 1 TAMPILAN
        # ===============================
        st.markdown("### 📊 Satker Terbaik & Terendah Berdasarkan Nilai IKPA")

        # =========================
        # BARIS 1 – TERBAIK
        # =========================
        c1, c2, c3 = st.columns(3)

        with c1:
            st.markdown(
                "<div style='margin-top:2px; margin-bottom:6px'><b>10 Satker Kecil Terbaik</b></div>",
                unsafe_allow_html=True
            )
            safe_chart(df_kecil, "KECIL", top=True, color="Greens",
                    y_min=y_min, y_max=y_max)

        with c2:
            st.markdown(
                "<div style='margin-top:2px; margin-bottom:6px'><b>10 Satker Sedang Terbaik</b></div>",
                unsafe_allow_html=True
            )
            safe_chart(df_sedang, "SEDANG", top=True, color="Greens",
                    y_min=y_min, y_max=y_max)

        with c3:
            st.markdown(
                "<div style='margin-top:2px; margin-bottom:6px'><b>10 Satker Besar Terbaik</b></div>",
                unsafe_allow_html=True
            )
            safe_chart(df_besar, "BESAR", top=True, color="Greens",
                    y_min=y_min, y_max=y_max)

        # ⬇️ JARAK ANTAR BARIS (rapat tapi aman)
        st.markdown("<div style='height:2px'></div>", unsafe_allow_html=True)

        # =========================
        # BARIS 2 – TERENDAH
        # =========================
        c4, c5, c6 = st.columns(3)

        with c4:
            st.markdown(
                "<div style='margin-top:2px; margin-bottom:6px'><b>10 Satker Kecil Terendah</b></div>",
                unsafe_allow_html=True
            )
            safe_chart(df_kecil, "KECIL", top=False, color="Reds",
                    y_min=y_min, y_max=y_max)

        with c5:
            st.markdown(
                "<div style='margin-top:2px; margin-bottom:6px'><b>10 Satker Sedang Terendah</b></div>",
                unsafe_allow_html=True
            )
            safe_chart(df_sedang, "SEDANG", top=False, color="Reds",
                    y_min=y_min, y_max=y_max)

        with c6:
            st.markdown(
                "<div style='margin-top:2px; margin-bottom:6px'><b>10 Satker Besar Terendah</b></div>",
                unsafe_allow_html=True
            )
            safe_chart(df_besar, "BESAR", top=False, color="Reds",
                    y_min=y_min, y_max=y_max)


        # Satker dengan masalah (Deviasi Hal 3 DIPA)
        st.subheader("🚨 Satker yang Memerlukan Perhatian Khusus")
        st.markdown("###### Atur Skala Nilai (Sumbu Y)")
        col_min_dev, col_max_dev = st.columns(2)
        with col_min_dev:
            y_min_dev = st.slider(
                "Nilai Minimum (Y-Axis)",
                min_value=0,
                max_value=50,
                value=40,
                step=1,
                key="high_ymin_dev"
            )
        with col_max_dev:
            y_max_dev = st.slider(
                "Nilai Maksimum (Y-Axis)",
                min_value=51,
                max_value=110,
                value=110,
                step=1,
                key="high_ymax_dev"
            )

        fig_dev = create_problem_chart(
            df, 
            'Deviasi Halaman III DIPA', 
            90, 
            "Deviasi Hal 3 DIPA Belum Optimal (< 90)",
            'less',
            y_min=y_min_dev,
            y_max=y_max_dev,
            show_yaxis=True
        )
        if fig_dev:
            st.plotly_chart(fig_dev, use_container_width=True)
        else:
            st.success("✅ Semua satker sudah optimal untuk Deviasi Hal 3 DIPA")

    # -------------------------
    # DATA DETAIL SATKER
    # -------------------------
    else:
        st.subheader("📋 Tabel Detail Satker")

        # persistent sub-tab for Periodik / Detail Satker
        if "active_table_tab" not in st.session_state:
            st.session_state.active_table_tab = "📆 Periodik"

        sub_tab = st.radio(
            "Pilih Mode Tabel",
            ["📆 Periodik", "📋 Detail Satker"],
            key="sub_tab_choice",
            horizontal=True
        )
        st.session_state['active_table_tab'] = sub_tab

        # -------------------------
        # PERIODIK TABLE
        # -------------------------
        if sub_tab == "📆 Periodik":
            st.markdown("#### Periodik — ringkasan per bulan / triwulan / perbandingan")

            # Tentukan tahun yang tersedia
            years = set()
            for k, df_period in st.session_state.data_storage.items():
                years.update(df_period['Tahun'].astype(str).unique())
            years = sorted([int(y) for y in years if str(y).strip() != ''], reverse=True)

            if not years:
                st.info("Tidak ada data periodik untuk ditampilkan.")
                st.stop()

            default_year = years[0]
            selected_year = st.selectbox("Pilih Tahun", options=years, index=0, key='tab_periodik_year_select')

            # session state untuk period_type
            if "period_type" not in st.session_state:
                st.session_state.period_type = "quarterly"

            period_options = ["quarterly", "monthly", "compare"]
            try:
                period_index = period_options.index(st.session_state.period_type)
            except ValueError:
                period_index = 0
                st.session_state.period_type = "quarterly"

            # Radio button
            period_type = st.radio(
                "Jenis Periode",
                options=period_options,
                format_func=lambda x: {"quarterly": "Triwulan", "monthly": "Bulanan", "compare": "Perbandingan"}.get(x, x),
                horizontal=True,
                index=period_index,
                key="period_type_radio_v2"
            )
            st.session_state.period_type = period_type

            # Pilih indikator (satu untuk semua mode)
            indicator_options = [
                'Kualitas Perencanaan Anggaran', 'Kualitas Pelaksanaan Anggaran', 'Kualitas Hasil Pelaksanaan Anggaran',
                'Revisi DIPA', 'Deviasi Halaman III DIPA', 'Penyerapan Anggaran', 'Belanja Kontraktual',
                'Penyelesaian Tagihan', 'Pengelolaan UP dan TUP', 'Capaian Output', 'Dispensasi SPM (Pengurang)',
                'Nilai Akhir (Nilai Total/Konversi Bobot)'
            ]
            default_indicator = 'Deviasi Halaman III DIPA'
            selected_indicator = st.selectbox(
                "Pilih Indikator", 
                options=indicator_options, 
                index=indicator_options.index(default_indicator) if default_indicator in indicator_options else 0,
                key='tab_periodik_indicator_select'
            )
            
            # -------------------------
            # Monthly / Quarterly
            # -------------------------
            if period_type in ['monthly', 'quarterly']:

                # 1. Gabungkan data per tahun
                dfs = []
                for (mon, yr), df_period in st.session_state.data_storage.items():
                    try:
                        if int(yr) == int(selected_year):
                            temp = df_period.copy()

                            # ambil kolom bulan apa pun namanya
                            if 'Bulan' in temp.columns:
                                temp['Bulan_raw'] = temp['Bulan']
                            elif 'Nama Bulan' in temp.columns:
                                temp['Bulan_raw'] = temp['Nama Bulan']
                            else:
                                continue

                            dfs.append(temp)
                    except:
                        continue

                if not dfs:
                    st.info(f"Tidak ditemukan data untuk tahun {selected_year}.")
                    st.stop()

                df_year = pd.concat(dfs, ignore_index=True)

                # =========================================================
                # 2. Normalisasi bulan (SUPER DEFENSIVE)
                # =========================================================
                MONTH_FIX = {
                    "JAN": "JANUARI", "JANUARI": "JANUARI",
                    "FEB": "FEBRUARI", "FEBRUARI": "FEBRUARI",
                    "MAR": "MARET", "MRT": "MARET", "MARET": "MARET",
                    "APR": "APRIL", "APRIL": "APRIL",
                    "MEI": "MEI",
                    "JUN": "JUNI", "JUNI": "JUNI",
                    "JUL": "JULI", "JULI": "JULI",
                    "AGT": "AGUSTUS", "AGUSTUS": "AGUSTUS",
                    "SEP": "SEPTEMBER", "SEPTEMBER": "SEPTEMBER",
                    "OKT": "OKTOBER", "OKTOBER": "OKTOBER",
                    "DES": "DESEMBER", "DESEMBER": "DESEMBER"
                }

                df_year['Bulan_upper'] = (
                    df_year['Bulan_raw']
                    .astype(str)
                    .str.upper()
                    .str.strip()
                    .map(lambda x: MONTH_FIX.get(x, x))
                )

                # =========================================================
                # 3. Period Column & Order (INI KUNCI)
                # =========================================================
                if period_type == 'monthly':
                    df_year['Period_Column'] = df_year['Bulan_upper']
                    df_year['Period_Order'] = df_year['Bulan_upper'].map(MONTH_ORDER)

                else:  # quarterly
                    def to_quarter(m):
                        return {
                        'JANUARI': 'Tw I', 'FEBRUARI': 'Tw I', 'MARET': 'Tw I',
                        'APRIL': 'Tw II', 'MEI': 'Tw II', 'JUNI': 'Tw II',
                        'JULI': 'Tw III', 'AGUSTUS': 'Tw III', 'SEPTEMBER': 'Tw III',
                        'OKTOBER': 'Tw IV', 'NOVEMBER': 'Tw IV', 'DESEMBER': 'Tw IV'
                    }.get(m)


                    quarter_order = {'Tw I':1,'Tw II':2,'Tw III':3,'Tw IV':4}
                    df_year['Period_Column'] = df_year['Bulan_upper'].map(to_quarter)
                    df_year['Period_Order'] = df_year['Period_Column'].map(quarter_order)

                # =========================================================
                # 4. PIVOT LANGSUNG (FIXED)
                # =========================================================

                # --- pilih nama SATKER TERPENDEK per Kode Satker ---
                name_map = (
                    df_year
                    .assign(name_len=df_year['Uraian Satker-RINGKAS'].astype(str).str.len())
                    .sort_values('name_len')
                    .groupby('Kode Satker')['Uraian Satker-RINGKAS']
                    .first()
                )

                df_pivot = df_year[
                    [
                        'Kode BA',
                        'Kode Satker',
                        'Period_Column',
                        selected_indicator
                    ]
                ].copy()

                df_wide = (
                    df_pivot
                    .pivot_table(
                        index=['Kode BA','Kode Satker'],  # ❗ IDENTIFIER ONLY
                        columns='Period_Column',
                        values=selected_indicator,
                        aggfunc='last'
                    )
                    .reset_index()
                )

                # --- pasang kembali nama satker ---
                df_wide['Uraian Satker-RINGKAS'] = df_wide['Kode Satker'].map(name_map)

                # =========================================================
                # 5. Urutkan kolom periode
                # =========================================================
                if period_type == 'monthly':
                    ordered_periods = sorted(
                        [c for c in df_wide.columns if c in MONTH_ORDER],
                        key=lambda x: MONTH_ORDER[x]
                    )
                else:
                    ordered_periods = [c for c in ['Tw I','Tw II','Tw III','Tw IV'] if c in df_wide.columns]

                # =========================================================
                # 6. Ranking 
                # =========================================================
                if ordered_periods:
                    last = ordered_periods[-1]
                    df_wide['Latest_Value'] = pd.to_numeric(df_wide[last], errors='coerce')
                    df_wide['Peringkat'] = (
                        df_wide['Latest_Value']
                        .rank(ascending=False, method='dense')
                        .astype('Int64')
                    )

                df_wide = df_wide.sort_values('Peringkat')

                # =========================================================
                # 7. DISPLAY 
                # =========================================================
                display_cols = ['Peringkat','Kode BA','Kode Satker','Uraian Satker-RINGKAS'] + ordered_periods
                df_display = df_wide[display_cols].copy()

                if period_type == 'monthly':
                    df_display.rename(columns={m: m.capitalize() for m in ordered_periods}, inplace=True)
                    display_period_cols = [m.capitalize() for m in ordered_periods]
                else:
                    display_period_cols = ordered_periods

                df_display[display_period_cols] = df_display[display_period_cols].fillna("–")

                # =============================
                # SEARCH & STYLING 
                # =============================
                search_query = st.text_input(
                    "🔎 Cari (Periodik) – ketik untuk filter di semua kolom",
                    value="",
                    key='tab_periodik_search'
                )

                if search_query:
                    q = str(search_query).strip().lower()
                    mask = df_display.apply(
                        lambda row: row.astype(str).str.lower().str.contains(q, na=False).any(),
                        axis=1
                    )
                    df_display_filtered = df_display[mask].copy()
                else:
                    df_display_filtered = df_display.copy()

                # Trend coloring
                def color_trend(row):
                    styles = []

                    # ambil hanya nilai numerik (buang "–", NaN, dll)
                    vals = []
                    for c in display_period_cols:
                        try:
                            v = float(row[c])
                            if not pd.isna(v):
                                vals.append(v)
                        except (ValueError, TypeError):
                            continue

                    # default: tidak ada warna
                    color = ''

                    if len(vals) >= 2:
                        if vals[-1] > vals[-2]:
                            color = 'background-color: #c6efce'  # hijau
                        elif vals[-1] < vals[-2]:
                            color = 'background-color: #f8d7da'  # merah

                    for c in row.index:
                        if display_period_cols and c == display_period_cols[-1]:
                            styles.append(color)
                        else:
                            styles.append('')

                    return styles

                def highlight_top(s):
                    if s.name == 'Peringkat':
                        return [
                            'background-color: gold' if (pd.to_numeric(v, errors='coerce') <= 3) else ''
                            for v in s
                        ]
                    return ['' for _ in s]

                styler = df_display_filtered.style.format(precision=2, na_rep='–')
                if display_period_cols:
                    styler = styler.apply(color_trend, axis=1)
                styler = styler.apply(highlight_top)

                st.dataframe(styler, use_container_width=True, height=600)


            elif period_type == "compare":
                st.markdown("### Perbandingan Antara Dua Tahun")

               # Gabungkan seluruh data
                all_data = []
                for (mon, yr), df in st.session_state.data_storage.items():
                    df2 = df.copy()
                    df2["Bulan_upper"] = df2["Bulan"].astype(str).str.upper().str.strip()
                    df2["Tahun"] = df2["Tahun"].astype(int)
                    all_data.append(df2)

                if not all_data:
                    st.warning("Belum ada data yang di-upload.")
                    st.stop()

                df_full = pd.concat(all_data, ignore_index=True)


                # Tahun yang valid
                available_years = sorted([y for y in df_full["Tahun"].unique() if 2022 <= y <= 2025])
                if len(available_years) < 2:
                    st.warning("Data tahun tidak cukup.")
                    st.stop()

                # Pilih Tahun A dan B
                colA, colB = st.columns(2)
                with colA:
                    year_a = st.selectbox("Tahun A (Awal)", available_years, index=0, key="tahunA_compare")
                with colB:
                    year_b = st.selectbox("Tahun B (Akhir)", available_years, index=1, key="tahunB_compare")

                if year_a == year_b:
                    st.info("Pilih dua tahun yang berbeda.")
                    st.stop()

                # Filter tahun
                df_a = df_full[df_full["Tahun"] == year_a]
                df_b = df_full[df_full["Tahun"] == year_b]

                def extract_tw(df_):
                    return {
                        "Tw I": df_[df_["Bulan_upper"] == "MARET"],
                        "Tw II": df_[df_["Bulan_upper"] == "JUNI"],
                        "Tw III": df_[df_["Bulan_upper"] == "SEPTEMBER"],
                        "Tw IV": df_[df_["Bulan_upper"] == "DESEMBER"],
                    }

                tw_a = extract_tw(df_a)
                tw_b = extract_tw(df_b)

                # Pilihan Satker
                satker_list = df_full[['Kode Satker', 'Uraian Satker-RINGKAS']].drop_duplicates()
                satker_options = ["SEMUA SATKER"] + satker_list['Kode Satker'].tolist()

                selected_satkers = st.multiselect(
                    "Pilih Satker",
                    satker_options,
                    format_func=lambda x: (
                        "SEMUA SATKER" if x == "SEMUA SATKER"
                        else satker_list[satker_list['Kode Satker'] == x]['Uraian Satker-RINGKAS'].values[0]
                    ),
                    default=["SEMUA SATKER"],
                    key="satker_compare"
                )

                if "SEMUA SATKER" in selected_satkers:
                    selected_satkers_final = satker_list['Kode Satker'].tolist()
                else:
                    selected_satkers_final = selected_satkers

                # Build tabel
                rows = []
                for _, m in satker_list.iterrows():
                    kode = m['Kode Satker']
                    if kode not in selected_satkers_final:
                        continue

                    nama = m['Uraian Satker-RINGKAS']
                    row = {"Kode Satker": kode, "Uraian Satker": nama}

                    latest_a = None
                    latest_b = None

                    for tw in ['Tw I', 'Tw II', 'Tw III', 'Tw IV']:
                        # Tahun A
                        valA = tw_a[tw][tw_a[tw]['Kode Satker'] == kode][selected_indicator].values
                        valA = valA[0] if len(valA) else None
                        row[f"{tw} {year_a}"] = valA
                        if valA is not None:
                            latest_a = valA

                        # Tahun B
                        valB = tw_b[tw][tw_b[tw]['Kode Satker'] == kode][selected_indicator].values
                        valB = valB[0] if len(valB) else None
                        row[f"{tw} {year_b}"] = valB
                        if valB is not None:
                            latest_b = valB

                    # Selisih
                    if latest_a is not None and latest_b is not None:
                        row[f"Δ Total ({year_b}-{year_a})"] = latest_b - latest_a
                    else:
                        row[f"Δ Total ({year_b}-{year_a})"] = None

                    rows.append(row)

                df_compare = pd.DataFrame(rows)

                # Styling warna
                def highlight_years(col_name):
                    if str(year_a) in col_name:
                        return 'background-color: #FFF8C6;'  # kuning muda
                    if str(year_b) in col_name:
                        return 'background-color: #DCEBFF;'  # biru muda
                    return ''

                df_style = df_compare.style.apply(
                    lambda row: [highlight_years(col) for col in df_compare.columns],
                    axis=1
                ).format(precision=2)

                st.markdown("### Hasil Perbandingan")
                st.dataframe(df_style, use_container_width=True, height=600)

        # -------------------------
        # DETAIL SATKER (legacy table)
        # -------------------------
        else:
            # ensure df available (use selected period if set)
            df = st.session_state.data_storage.get(st.session_state.get('selected_period', all_periods[0]), None)
            if df is None:
                st.info("Data untuk detail satker tidak tersedia untuk periode yang dipilih.")
                return

            col1, col2 = st.columns([2, 1])
            with col1:
                view_mode = st.radio(
                    "Tampilan",
                    options=['aspek', 'komponen'],
                    format_func=lambda x: 'Berdasarkan Aspek' if x == 'aspek' else 'Berdasarkan Komponen',
                    horizontal=True
                )
            with col2:
                st.write("")

            display_columns = ['Peringkat', 'Kode BA', 'Kode Satker', 'Uraian Satker-RINGKAS']
            if view_mode == 'aspek':
                display_columns += [
                    'Kualitas Perencanaan Anggaran',
                    'Kualitas Pelaksanaan Anggaran',
                    'Kualitas Hasil Pelaksanaan Anggaran'
                ]
                df_display = df[display_columns + ['Nilai Total',
                                                   'Dispensasi SPM (Pengurang)',
                                                   'Nilai Akhir (Nilai Total/Konversi Bobot)']].copy()
            else:
                component_cols = [
                    'Revisi DIPA', 'Deviasi Halaman III DIPA', 'Penyerapan Anggaran',
                    'Belanja Kontraktual', 'Penyelesaian Tagihan', 
                    'Pengelolaan UP dan TUP', 'Capaian Output'
                ]
                df_display = df[display_columns + ['Nilai Total',
                                                   'Dispensasi SPM (Pengurang)',
                                                   'Nilai Akhir (Nilai Total/Konversi Bobot)']].copy()
                for col in component_cols:
                    df_display[col] = df.get(col, 0)
                final_cols = display_columns + component_cols + ['Nilai Total',
                                                                 'Dispensasi SPM (Pengurang)',
                                                                 'Nilai Akhir (Nilai Total/Konversi Bobot)']
                df_display = df_display[final_cols]

            # Search widget & styling
            search_query = st.text_input("🔎 Cari (ketik untuk filter di semua kolom)", value="", help="Cari teks pada semua kolom (case-insensitive).", key='search_detail')
            if search_query:
                q = str(search_query).strip().lower()
                mask = df_display.apply(lambda row: row.astype(str).str.lower().str.contains(q, na=False).any(), axis=1)
                df_display_filtered = df_display[mask].copy()
            else:
                df_display_filtered = df_display.copy()

            def highlight_top(s):
                if s.name == 'Peringkat':
                    return ['background-color: gold' if (pd.to_numeric(v, errors='coerce') <= 3) else '' for v in s]
                return ['' for _ in s]

            st.dataframe(
                df_display_filtered.style.apply(highlight_top).format(precision=2),
                use_container_width=True,
                height=600
            )


# HALAMAN 2: DASHBOARD INTERNAL KPPN (Protected)    
def menu_ews_satker():
    st.subheader("🏛️ Early Warning System Kinerja Keuangan Satker")

    if "data_storage" not in st.session_state or not st.session_state.data_storage:
        st.warning("⚠️ Belum ada data historis yang tersedia.")
        return
    
    # Gabungkan semua data
    all_data = []
    for period, df in st.session_state.data_storage.items():
        df_copy = df.copy()
        # ensure Period & Period_Sort exist
        df_copy['Period'] = f"{period[0]} {period[1]}"
        df_copy['Period_Sort'] = f"{period[1]}-{period[0]}"
        all_data.append(df_copy)
    
    if not all_data:
        st.warning("⚠️ Belum ada data historis yang tersedia.")
        return
    
    df_all = pd.concat(all_data, ignore_index=True)
      
    # Analisis tren dan Early Warning System
    # Gunakan data periode terkini
    latest_period = sorted(st.session_state.data_storage.keys(), key=lambda x: (int(x[1]), MONTH_ORDER.get(x[0].upper(), 0)), reverse=True)[0]
    df_latest = st.session_state.data_storage[latest_period]

    st.markdown("---")
    st.subheader("🚨 Satker yang Memerlukan Perhatian Khusus")

    # 🎚️ Pengaturan Sumbu Y
    st.markdown("###### Atur Skala Nilai (Sumbu Y)")
    col_min, col_max = st.columns(2)
    with col_min:
        y_min_int = st.slider(
            "Nilai Minimum (Y-Axis)",
            min_value=0,
            max_value=50,
            value=50,
            step=1,
            key="ymin_internal"
        )
    with col_max:
        y_max_int = st.slider(
            "Nilai Maksimum (Y-Axis)",
            min_value=51,
            max_value=110,
            value=110,
            step=1,
            key="ymax_internal"
        )

    # 📊 Highlights Kinerja Satker yang Perlu Perhatian Khusus
    col1, col2 = st.columns(2)

    with col1:
        fig_up = create_problem_chart(
            df_latest,
            'Pengelolaan UP dan TUP',
            100,
            "Pengelolaan UP dan TUP Belum Optimal (< 100)",
            'less',
            y_min=y_min_int,
            y_max=y_max_int,
            show_yaxis=True  # Left chart shows Y-axis
        )
        if fig_up:
            st.plotly_chart(fig_up, use_container_width=True)
        else:
            st.success("✅ Semua satker sudah optimal untuk Pengelolaan UP dan TUP")

    with col2:
        fig_output = create_problem_chart(
            df_latest,
            'Capaian Output',
            100,
            "Capaian Output Belum Optimal (< 100)",
            'less',
            y_min=y_min_int,
            y_max=y_max_int,
            show_yaxis=False  # Right chart hides Y-axis
        )
        if fig_output:
            st.plotly_chart(fig_output, use_container_width=True)
        else:
            st.success("✅ Semua satker sudah optimal untuk Capaian Output")
    
    warnings = []

    st.markdown("---")
    
    # Analisis Tren
    st.subheader("📈 Analisis Tren")
    
    col1, col2, col3 = st.columns(3)
    
    with col1:
        # 🔍 DETAILED ERROR CHECKING
        # Map month names to numbers
        df_all['Month_Num'] = df_all['Bulan'].str.strip().str.upper().map(MONTH_ORDER)
        
        # Check for unmapped months
        missing_months = df_all[df_all['Month_Num'].isna()]
        if len(missing_months) > 0:
            st.error("❌ **DITEMUKAN BULAN YANG TIDAK VALID:**")
            
            # Group by period to show which files have issues
            problem_periods = missing_months.groupby(['Bulan', 'Tahun']).size().reset_index(name='Count')
            
            for _, row in problem_periods.iterrows():
                st.warning(f"⚠️ Periode **{row['Bulan']} {row['Tahun']}** - Nama bulan '{row['Bulan']}' tidak dikenali (ditemukan di {row['Count']} baris)")
            
            st.info("""
            **Solusi:**
            1. Periksa file Excel untuk periode yang bermasalah
            2. Pastikan nama bulan sesuai format: JANUARI, FEBRUARI, MARET, dst (huruf besar)
            3. Upload ulang file yang bermasalah dari halaman Admin
            """)
            
            # Show expected month names
            with st.expander("📋 Lihat format bulan yang valid"):
                st.write("Format yang diterima:")
                st.code(", ".join(MONTH_ORDER.keys()))
            
            # Option to proceed with cleaned data
            if st.checkbox("⚠️ Abaikan data bermasalah dan lanjutkan"):
                df_all = df_all.dropna(subset=['Month_Num'])
                st.info(f"✅ Data dibersihkan. Sisa {len(df_all)} baris.")
            else:
                st.stop()
        
        # Check for invalid years
        invalid_years = df_all[df_all['Tahun'].isna()]
        if len(invalid_years) > 0:
            st.error("❌ **DITEMUKAN TAHUN YANG TIDAK VALID:**")
            
            problem_periods = invalid_years.groupby(['Bulan']).size().reset_index(name='Count')
            for _, row in problem_periods.iterrows():
                st.warning(f"⚠️ Bulan **{row['Bulan']}** - Tahun tidak valid (ditemukan di {row['Count']} baris)")
            
            st.stop()
        
        # Try to create Period_Sort with detailed error handling
        try:
            # Convert to int safely
            df_all['Tahun_Int'] = df_all['Tahun'].astype(int)
            df_all['Month_Num_Int'] = df_all['Month_Num'].astype(int)
            
            # Create Period_Sort
            df_all['Period_Sort'] = df_all.apply(
                lambda x: f"{x['Tahun_Int']:04d}-{x['Month_Num_Int']:02d}", 
                axis=1
            )
                        
        except Exception as e:
            st.error(f"❌ **ERROR saat membuat Period_Sort:** {str(e)}")
            
            # Show problematic rows
            st.write("**Baris yang bermasalah:**")
            problem_cols = ['Bulan', 'Tahun', 'Month_Num', 'Kode Satker', 'Uraian Satker']
            st.dataframe(df_all[problem_cols].head(20))
            
            st.stop()
        
        # Now create the selectbox
        available_periods = sorted(df_all['Period_Sort'].unique())
        start_period = st.selectbox(
            "Periode Awal",
            options=available_periods,
            index=0
        )
    
    with col2:
        end_period = st.selectbox(
            "Periode Akhir",
            options=available_periods,
            index=len(available_periods) - 1
        )
    
    # Filter berdasarkan periode
    df_filtered = df_all[
        (df_all['Period_Sort'] >= start_period) & 
        (df_all['Period_Sort'] <= end_period)
    ]
    
    with col3:
        # Pilihan metrik
        metric_options = {
            'Nilai Akhir (Nilai Total/Konversi Bobot)': 'Nilai Akhir (Nilai Total/Konversi Bobot)',
            'Kualitas Perencanaan Anggaran': 'Kualitas Perencanaan Anggaran',
            'Kualitas Pelaksanaan Anggaran': 'Kualitas Pelaksanaan Anggaran',
            'Kualitas Hasil Pelaksanaan Anggaran': 'Kualitas Hasil Pelaksanaan Anggaran',
            'Revisi DIPA': 'Revisi DIPA',
            'Deviasi Halaman III DIPA': 'Deviasi Halaman III DIPA',
            'Penyerapan Anggaran': 'Penyerapan Anggaran',
            'Belanja Kontraktual': 'Belanja Kontraktual',
            'Penyelesaian Tagihan': 'Penyelesaian Tagihan',
            'Pengelolaan UP dan TUP': 'Pengelolaan UP dan TUP',
            'Capaian Output': 'Capaian Output'
        }
        
        selected_metric = st.selectbox(
            "Metrik yang Ditampilkan",
            options=list(metric_options.keys()),
            index=0
        )
    
    # Pilih satker
    #  AMBIL PERIODE TERBARU (AMAN)
    def period_sort_key(k):
        mon, yr = k
        try:
            y = int(yr)
        except Exception as e:
            st.warning(f"⚠️ Tidak bisa convert tahun '{yr}' untuk periode {mon}: {e}")
            y = 0
        return (y, MONTH_ORDER.get(mon.upper(), 0))

    try:
        latest_period = sorted(
            st.session_state.data_storage.keys(),
            key=period_sort_key,
            reverse=True
        )[0]
        latest_df = st.session_state.data_storage[latest_period].copy()
    except Exception as e:
        st.error(f"❌ Error mendapatkan periode terbaru: {e}")
        st.write("**Periode yang tersedia:**")
        st.write(list(st.session_state.data_storage.keys()))
        st.stop()

    # ======================================================
    # 📌 PASTIKAN KOLOM KODE SATKER
    # ======================================================
    if 'Kode Satker' in latest_df.columns:
        latest_df['Kode Satker'] = latest_df['Kode Satker'].astype(str)
    else:
        latest_df['Kode Satker'] = latest_df.index.astype(str)

    bottom_10_default = (
        latest_df
        .nsmallest(10, 'Nilai Akhir (Nilai Total/Konversi Bobot)')['Kode Satker']
        .astype(str)
        .tolist()
    )

    # ======================================================
    # 📌 PASTIKAN KOLOM SATKER (AMAN UNTUK SEMUA KONDISI)
    # ======================================================
    if "Satker" not in df_all.columns:
        if "Uraian Satker" in df_all.columns and "Kode Satker" in df_all.columns:
            df_all["Satker"] = (
                df_all["Uraian Satker"].astype(str)
                + " ("
                + df_all["Kode Satker"].astype(str)
                + ")"
            )
        elif "Uraian Satker" in df_all.columns:
            df_all["Satker"] = df_all["Uraian Satker"].astype(str)
        else:
            st.error("❌ Kolom Satker tidak tersedia pada data.")
            st.stop()

    # ======================================================
    # 📌 FILTER PERIODE (PASTIKAN df_filtered ADA)
    # ======================================================
    df_filtered = df_all.copy()

    # ======================================================
    # 📌 PILIH SATKER
    # ======================================================
    all_satker = sorted(df_filtered["Satker"].dropna().unique())

    selected_satker = st.multiselect(
        "Pilih Satker",
        options=all_satker,
        default=[
            s for s in all_satker
            if any(code in s for code in bottom_10_default)
        ][:10]
    )

    if not selected_satker:
        st.warning("Silakan pilih minimal satu satker untuk melihat tren.")
        return

    # ======================================================
    # 📌 DATA UNTUK PLOT
    # ======================================================
    df_plot = df_filtered[df_filtered["Satker"].isin(selected_satker)]

    
    # Buat line chart
    fig = go.Figure()
    
    try:
        for satker in selected_satker:
            df_satker = df_plot[df_plot['Satker'] == satker].sort_values('Period_Sort')

            # Ensure x-axis uses correct chronological month order
            categories = [f"{m} {y}" for y, m in sorted(
                {(int(x['Tahun']), x['Bulan'].upper()) for _, x in df_all.iterrows()},
                key=lambda t: (t[0], MONTH_ORDER.get(t[1], 0))
            )]
            
            fig.add_trace(go.Scatter(
                x=pd.Categorical(
                    df_satker['Period'],
                    categories=categories,
                    ordered=True
                ),
                y=df_satker[selected_metric],
                mode='lines+markers',
                name=satker,
                hovertemplate='<b>%{fullData.name}</b><br>Periode: %{x}<br>Nilai: %{y:.2f}<extra></extra>'
            ))
    except Exception as e:
        st.error(f"❌ Error membuat chart: {str(e)}")
        st.write("**Debug Info:**")
        st.write(f"Selected satker: {selected_satker}")
        st.write(f"df_plot shape: {df_plot.shape}")
        st.write(f"Unique periods in df_plot: {df_plot['Period'].unique()}")
        st.stop()
    
    fig.update_layout(
        title=f"Tren {selected_metric}",
        xaxis_title="Periode",
        yaxis_title="Nilai",
        height=600,
        hovermode='x unified',
        legend=dict(
            orientation="v",
            yanchor="top",
            y=1,
            xanchor="left",
            x=1.02
        )
    )
    
    st.plotly_chart(fig, use_container_width=True)

    # Early Warning Satker Tren Menurun
    warnings = []  # Initialize warnings list
    
    for satker in selected_satker:
        df_satker = df_plot[df_plot['Satker'] == satker].sort_values('Period_Sort')
        
        if len(df_satker) >= 2:
            values = df_satker[selected_metric].values
            
            # Cek tren menurun (2 periode terakhir)
            if len(values) >= 2:
                last_value = values[-1]
                prev_value = values[-2]
                
                if last_value < prev_value:
                    decrease = prev_value - last_value
                    warnings.append({
                        'Satker': satker,
                        'Metrik': selected_metric,
                        'Nilai Sebelumnya': prev_value,
                        'Nilai Terkini': last_value,
                        'Penurunan': decrease
                    })
    
    if warnings:
        st.warning(f"⚠️ Ditemukan {len(warnings)} satker dengan tren menurun!")
        
        for w in warnings:
            st.markdown(f"""
            **{w['Satker']}**  
            - Metrik: {w['Metrik']}
            - Nilai sebelumnya: {w['Nilai Sebelumnya']:.2f}
            - Nilai terkini: {w['Nilai Terkini']:.2f}
            - Penurunan: {w['Penurunan']:.2f} poin
            """)
            st.markdown("---")
    else:
        st.success("✅ Tidak ada satker dengan tren menurun pada periode yang dipilih!")
        
#HIGHLIGHTS
def menu_highlights():
    st.subheader("🎯 Highlights IKPA KPPN")

    # ===============================
    # VALIDASI DATA
    # ===============================
    if "data_storage_kppn" not in st.session_state or not st.session_state.data_storage_kppn:
        st.info("ℹ️ Belum ada data IKPA KPPN yang tersimpan.")
        return

    # ===============================
    # GABUNGKAN DATA IKPA KPPN
    # ===============================
    all_data = []
    for (bulan, tahun), df in st.session_state.data_storage_kppn.items():
        df_copy = df.copy()
        df_copy["Periode"] = f"{bulan} {tahun}"
        df_copy["Tahun"] = int(tahun)
        df_copy["Bulan"] = bulan
        all_data.append(df_copy)

    df_all = pd.concat(all_data, ignore_index=True)

    # ===============================
    # NORMALISASI KOLOM
    # ===============================
    df_all.columns = (
        df_all.columns.astype(str)
        .str.replace(r"\s+", " ", regex=True)
        .str.strip()
    )

    # ===============================
    # PERBAIKI UNNAMED (IKPA)
    # ===============================
    rename_map = {
        "Unnamed: 5": "Revisi DIPA",
        "Unnamed: 7": "Deviasi Halaman III DIPA",
        "Unnamed: 8": "Penyerapan Anggaran",
        "Unnamed: 9": "Belanja Kontraktual",
        "Unnamed: 10": "Penyelesaian Tagihan",
        "Unnamed: 11": "Pengelolaan UP dan TUP",
        "Unnamed: 12": "Capaian Output",
    }
    df_all = df_all.rename(columns=rename_map)

    # ===============================
    # 🔑 PASTIKAN PERIOD_SORT ADA
    # ===============================
    if "Period_Sort" not in df_all.columns:
        df_all["Month_Num"] = df_all["Bulan"].str.upper().map(MONTH_ORDER)
        df_all["Period_Sort"] = (
            df_all["Tahun"].astype(str)
            + "-"
            + df_all["Month_Num"].astype(int).astype(str).str.zfill(2)
        )

    # ===============================
    # 📅 FILTER PERIODE (BARU)
    # ===============================
    st.markdown("### 📅 Filter Periode")

    available_periods = sorted(df_all["Period_Sort"].dropna().unique())

    col1, col2 = st.columns(2)

    with col1:
        start_period = st.selectbox(
            "Periode Awal",
            options=available_periods,
            index=0
        )

    with col2:
        end_period = st.selectbox(
            "Periode Akhir",
            options=available_periods,
            index=len(available_periods) - 1
        )

    df_all = df_all[
        (df_all["Period_Sort"] >= start_period) &
        (df_all["Period_Sort"] <= end_period)
    ]

    if df_all.empty:
        st.warning("⚠️ Data kosong pada rentang periode tersebut.")
        return

    st.success(f"Data IKPA KPPN dimuat ({len(df_all)} baris)")

    # ===============================
    # PILIH KPPN
    # ===============================
    kppn_list = sorted(df_all["Nama KPPN"].dropna().unique())
    selected_kppn = st.selectbox("Pilih KPPN", kppn_list)

    df_kppn = df_all[df_all["Nama KPPN"] == selected_kppn].copy()

    # ===============================
    # FILTER BARIS NILAI
    # ===============================
    if "Keterangan" in df_kppn.columns:
        df_kppn = df_kppn[
            df_kppn["Keterangan"].astype(str).str.upper() == "NILAI"
        ]

    # ===============================
    # INDIKATOR
    # ===============================
    indikator_opsi = [
        "Kualitas Perencanaan Anggaran",
        "Revisi DIPA",
        "Deviasi Halaman III DIPA",
        "Penyerapan Anggaran",
        "Belanja Kontraktual",
        "Penyelesaian Tagihan",
        "Pengelolaan UP dan TUP",
        "Capaian Output",
        "Nilai Total",
        "Nilai Akhir (Nilai Total/Konversi Bobot)"
    ]

    for col in indikator_opsi:
        if col in df_kppn.columns:
            df_kppn[col] = (
                df_kppn[col]
                .astype(str)
                .str.replace("%", "", regex=False)
                .str.replace(",", ".", regex=False)
            )
            df_kppn[col] = pd.to_numeric(df_kppn[col], errors="coerce")

    selected_indikator = st.multiselect(
        "Pilih Indikator IKPA KPPN !",
        [c for c in indikator_opsi if c in df_kppn.columns],
        default=["Nilai Akhir (Nilai Total/Konversi Bobot)"]
    )

    if not selected_indikator:
        st.warning("⚠️ Pilih minimal satu indikator.")
        return

    # ===============================
    # URUT PERIODE
    # ===============================
    df_kppn = df_kppn.sort_values("Period_Sort")

    # ===============================
    # LINE CHART
    # ===============================
    fig = go.Figure()

    for indikator in selected_indikator:
        fig.add_trace(
            go.Scatter(
                x=df_kppn["Periode"],
                y=df_kppn[indikator],
                mode="lines+markers",
                name=indikator,
            )
        )

    fig.update_layout(
        title=f"📈 Tren IKPA KPPN – {selected_kppn}",
        xaxis_title="Periode",
        yaxis_title="Nilai",
        height=600,
        hovermode="x unified"
    )

    st.plotly_chart(fig, use_container_width=True)


def page_trend():
    st.title("📈 Dashboard Internal KPPN")

    # ===============================
    # AUTHENTICATION
    # ===============================
    if "authenticated" not in st.session_state:
        st.session_state.authenticated = False

    if not st.session_state.authenticated:
        st.warning("🔒 Halaman ini memerlukan autentikasi Admin.")
        password = st.text_input("Masukkan Password", type="password")

        if st.button("Login"):
            if password == ADMIN_PASSWORD:
                st.session_state.authenticated = True
                st.success("Login berhasil!")
                st.rerun()
            else:
                st.error("Password salah!")
        return

    # ===============================
    # MENU DASHBOARD INTERNAL
    # ===============================
    menu = st.radio(
        "Pilih Menu",
        [
            "🏛️ Early Warning System Kinerja Keuangan Satker",
            "🎯 Highlights"
        ],
        horizontal=True
    )

    st.markdown("---")

    # ===============================
    # 🔽 PANGGIL ISI MENU
    # ===============================
    if menu == "🏛️ Early Warning System Kinerja Keuangan Satker":
        menu_ews_satker()

    elif menu == "🎯 Highlights":
        menu_highlights()
   
# ============================================================
# 🔐 HALAMAN 3: ADMIN 
# ============================================================
# ======================================================================================
# PROCESS IKPA
# ======================================================================================
# ======================================================================================
# DETECT DIPA HEADER (ROBUST VERSION)
# ======================================================================================
def detect_dipa_header(uploaded_file):
    """
    Auto-detect header row dalam file DIPA mentah.
    Returns: DataFrame dengan header yang sudah benar
    """
    try:
        uploaded_file.seek(0)
        
        # Baca 20 baris pertama untuk preview
        preview = pd.read_excel(uploaded_file, header=None, nrows=20, dtype=str)
        
        # Keywords yang PASTI ada di header DIPA
        header_keywords = [
            "satker", "kode", "pagu", "jumlah", "dipa", 
            "tanggal", "revisi", "no", "status"
        ]
        
        header_row = None
        max_matches = 0
        
        # Cari baris dengan keyword terbanyak
        for i in range(len(preview)):
            row_text = " ".join(preview.iloc[i].fillna("").astype(str).str.lower())
            matches = sum(1 for kw in header_keywords if kw in row_text)
            
            if matches > max_matches:
                max_matches = matches
                header_row = i
        
        # Jika tidak ada yang cocok, gunakan baris 0
        if header_row is None or max_matches < 3:
            st.warning("⚠️ Header otomatis tidak terdeteksi, menggunakan baris pertama")
            header_row = 0
        else:
            st.info(f"✅ Header terdeteksi di baris {header_row + 1} (keyword match: {max_matches})")
        
        # Baca ulang dengan header yang benar
        uploaded_file.seek(0)
        df = pd.read_excel(uploaded_file, header=header_row, dtype=str)
        
        # Bersihkan nama kolom
        df.columns = (
            df.columns.astype(str)
            .str.replace("\n", " ", regex=False)
            .str.replace("\r", " ", regex=False)
            .str.replace("\s+", " ", regex=True)
            .str.strip()
            .str.upper()  # Normalize to uppercase
        )
        
        # Hapus baris kosong
        df = df.dropna(how='all').reset_index(drop=True)
        
        # Debug: tampilkan preview
        st.write("**Preview 5 baris pertama setelah deteksi header:**")
        st.dataframe(df.head(5))
        
        return df
        
    except Exception as e:
        st.error(f"❌ Error deteksi header: {e}")
        uploaded_file.seek(0)
        return pd.read_excel(uploaded_file, dtype=str)


# ======================================================================================
# CLEAN DIPA (SUPER ROBUST VERSION)
# ======================================================================================
def clean_dipa(df_raw):
    """
    Membersihkan file DIPA mentah dan mengembalikan format standar.
    """
    
    df = df_raw.copy()
    
    # Hapus kolom Unnamed
    df = df.loc[:, ~df.columns.astype(str).str.contains("Unnamed", case=False)]
    
    # Normalize column names untuk matching yang lebih mudah
    df.columns = df.columns.astype(str).str.upper().str.strip()
    
    st.write("**Kolom yang terdeteksi di file DIPA:**")
    st.write(list(df.columns))
    
    # ====== HELPER: Flexible column finder ======
    def find_col(keywords_list):
        """Find column by multiple possible names"""
        for col in df.columns:
            col_clean = str(col).upper().replace(" ", "").replace("_", "")
            for kw in keywords_list:
                kw_clean = kw.upper().replace(" ", "").replace("_", "")
                if kw_clean in col_clean:
                    return col
        return None
    
    # ====== 1. KODE SATKER & NAMA SATKER ======
    satker_col = find_col([
        "SATKER", "KODESATKER", "KODE SATKER", "KODE SATUAN KERJA",
        "SATUAN KERJA", "SATKER/KPPN"
    ])
    
    if satker_col is None:
        st.error("❌ Kolom Satker tidak ditemukan!")
        st.write("Kolom available:", list(df.columns))
        raise ValueError("Kolom Satker tidak ditemukan")
    
    st.success(f"✅ Kolom Satker ditemukan: {satker_col}")
    
    # Extract kode 6 digit
    df["Kode Satker"] = (
        df[satker_col].astype(str)
        .str.extract(r"(\d{6})", expand=False)
        .fillna("")
    )
    
    # Extract nama satker (remove kode di depan)
    df["Satker"] = (
        df[satker_col].astype(str)
        .str.replace(r"^\d{6}\s*-?\s*", "", regex=True)
        .str.strip()
    )
    
    # Jika nama kosong, gunakan format default
    mask_empty = (df["Satker"] == "") | (df["Satker"].isna())
    df.loc[mask_empty, "Satker"] = df.loc[mask_empty, "Kode Satker"] + " - SATKER"
    
    # ====== 2. TOTAL PAGU ======
    pagu_col = find_col([
        "PAGU", "JUMLAH", "TOTAL PAGU", "TOTALPAGU", "NILAI PAGU",
        "PAGU DIPA", "JUMLAH DIPA"
    ])
    
    if pagu_col:
        st.success(f"✅ Kolom Pagu ditemukan: {pagu_col}")
        df["Total Pagu"] = pd.to_numeric(df[pagu_col], errors="coerce").fillna(0).astype(int)
    else:
        st.warning("⚠️ Kolom Pagu tidak ditemukan, menggunakan 0")
        df["Total Pagu"] = 0
    
    # ====== 3. TANGGAL POSTING REVISI ======
    tgl_col = find_col([
        "TANGGAL POSTING", "TGL POSTING", "TANGGALPOSTING",
        "TANGGAL REVISI", "TGL REVISI", "TANGGAL", "DATE"
    ])
    
    if tgl_col:
        st.success(f"✅ Kolom Tanggal ditemukan: {tgl_col}")
        df["Tanggal Posting Revisi"] = pd.to_datetime(df[tgl_col], errors="coerce")
    else:
        st.warning("⚠️ Kolom Tanggal tidak ditemukan")
        df["Tanggal Posting Revisi"] = pd.NaT
    
    # Extract Tahun
    df["Tahun"] = df["Tanggal Posting Revisi"].dt.year
    df["Tahun"] = df["Tahun"].fillna(datetime.now().year).astype(int)
    
    # ====== 4. KOLOM LAINNYA (OPTIONAL) ======
    
    # NO
    no_col = find_col(["NO", "NOMOR", "NO."])
    df["NO"] = df[no_col].astype(str).str.strip() if no_col else ""
    
    # KEMENTERIAN
    kementerian_col = find_col(["KEMENTERIAN", "KEMENTRIAN", "KL", "K/L", "BA"])
    df["Kementerian"] = df[kementerian_col].astype(str).str.strip() if kementerian_col else ""
    
    # KODE STATUS HISTORY
    status_col = find_col(["STATUS HISTORY", "KODE STATUS", "KODESTATUS", "STATUS"])
    df["Kode Status History"] = df[status_col].astype(str).str.strip() if status_col else ""
    
    # JENIS REVISI
    jenis_col = find_col(["JENIS REVISI", "JENISREVISI", "TIPE REVISI"])
    df["Jenis Revisi"] = df[jenis_col].astype(str).str.strip() if jenis_col else ""
    
    # REVISI KE-
    revisi_col = find_col(["REVISI KE", "REVISIKE", "REVISI"])
    if revisi_col:
        df["Revisi ke-"] = pd.to_numeric(df[revisi_col], errors="coerce").fillna(0).astype(int)
    else:
        df["Revisi ke-"] = 0
    
    # NO DIPA
    nodipa_col = find_col(["NO DIPA", "NODIPA", "NOMOR DIPA", "NO. DIPA"])
    df["No Dipa"] = df[nodipa_col].astype(str).str.strip() if nodipa_col else ""
    
    # TANGGAL DIPA
    tgldipa_col = find_col(["TANGGAL DIPA", "TGL DIPA", "TGLDIPA"])
    if tgldipa_col:
        df["Tanggal Dipa"] = pd.to_datetime(df[tgldipa_col], errors="coerce").dt.strftime("%d-%m-%Y")
    else:
        df["Tanggal Dipa"] = ""
    
    # OWNER
    owner_col = find_col(["OWNER", "PEMILIK"])
    df["Owner"] = df[owner_col].astype(str).str.strip() if owner_col else ""
    
    # DIGITAL STAMP
    stamp_col = find_col(["DIGITAL STAMP", "DIGITALSTAMP", "STAMP", "TTD DIGITAL"])
    df["Digital Stamp"] = df[stamp_col].astype(str).str.strip() if stamp_col else ""
    
    # ====== FINAL: Susun kolom sesuai urutan ======
    final_columns = [
        "Kode Satker", "Satker", "Tahun", "Tanggal Posting Revisi", "Total Pagu",
        "NO", "Kementerian", "Kode Status History", "Jenis Revisi", "Revisi ke-",
        "No Dipa", "Tanggal Dipa", "Owner", "Digital Stamp"
    ]
    
    df_clean = df[final_columns].copy()
    
    # Ambil revisi terakhir per satker per tahun
    df_clean = df_clean.sort_values(["Kode Satker", "Tahun", "Tanggal Posting Revisi"])
    df_clean = df_clean.groupby(["Kode Satker", "Tahun"], as_index=False).tail(1)
    
    # Filter hanya kode satker yang valid (6 digit)
    df_clean = df_clean[df_clean["Kode Satker"].str.len() == 6]
    
    st.write(f"**Hasil cleaning: {len(df_clean)} baris satker valid**")
    
    return df_clean


# ======================================================================================
# ASSIGN JENIS SATKER
# ======================================================================================
def assign_jenis_satker(df):
    """Klasifikasi satker berdasarkan Total Pagu"""
    
    if df.empty or "Total Pagu" not in df.columns:
        df["Jenis Satker"] = "Satker Kecil"
        return df
    
    q70 = df["Total Pagu"].quantile(0.70)
    q40 = df["Total Pagu"].quantile(0.40)
    
    def classify(pagu):
        if pagu >= q70: return "Satker Besar"
        elif pagu >= q40: return "Satker Sedang"
        else: return "Satker Kecil"
    
    df["Jenis Satker"] = df["Total Pagu"].apply(classify)
    
    # Reorder: Jenis Satker setelah Total Pagu
    cols = list(df.columns)
    if "Jenis Satker" in cols and "Total Pagu" in cols:
        cols.remove("Jenis Satker")
        pagu_idx = cols.index("Total Pagu")
        cols.insert(pagu_idx + 1, "Jenis Satker")
        df = df[cols]
    
    return df


# ======================================================================================
# PROCESS UPLOADED DIPA (MAIN FUNCTION)
# ======================================================================================
def process_uploaded_dipa(uploaded_file, save_file_to_github):
    """Process file DIPA upload user dengan validasi ketat"""
    
    try:
        st.info("📄 Memulai proses upload DIPA...")

        # 1️⃣ Baca raw excel
        with st.spinner("Membaca file..."):
            raw = pd.read_excel(uploaded_file, header=None, dtype=str)

        if raw.empty:
            return None, None, "❌ File kosong"

        # 2️⃣ Standarisasi format
        with st.spinner("Menstandarisasi format DIPA..."):
            df_std = standardize_dipa(raw)

        if df_std.empty:
            return None, None, "❌ Data tidak berhasil distandarisasi atau tidak ada data valid"

        # 3️⃣ Validasi Tahun
        if "Tahun" not in df_std.columns or df_std["Tahun"].isna().all():
            st.warning("⚠️ Tahun tidak terdeteksi, menggunakan tahun sekarang")
            tahun_dipa = datetime.now().year
            df_std["Tahun"] = tahun_dipa
        else:
            tahun_dipa = int(df_std["Tahun"].mode()[0])
            df_std["Tahun"] = df_std["Tahun"].fillna(tahun_dipa)

        # 4️⃣ Validasi data
        st.write(f"**Validasi:** {len(df_std)} baris data valid terdeteksi")
        st.write(f"**Tahun:** {tahun_dipa}")
        st.write(f"**Rentang Pagu:** Rp {df_std['Total Pagu'].min():,.0f} - Rp {df_std['Total Pagu'].max():,.0f}")

        # 5️⃣ Normalisasi kode satker
        df_std["Kode Satker"] = df_std["Kode Satker"].apply(normalize_kode_satker)

        # 6️⃣ Merge dengan referensi (jika ada)
        if "reference_df" in st.session_state and not st.session_state.reference_df.empty:
            with st.spinner("Menggabungkan dengan data referensi..."):
                ref = st.session_state.reference_df.copy()
                ref["Kode Satker"] = ref["Kode Satker"].apply(normalize_kode_satker)

                df_std = df_std.merge(
                    ref[["Kode BA", "K/L", "Kode Satker"]],
                    on="Kode Satker",
                    how="left"
                )

                if "Kementerian" in df_std.columns and "K/L" in df_std.columns:
                    df_std["Kementerian"] = df_std["Kementerian"].fillna(df_std["K/L"])

        # 7️⃣ Klasifikasi Satker
        with st.spinner("Mengklasifikasi jenis satker..."):
            df_std = assign_jenis_satker(df_std)

        # 8️⃣ Ambil revisi terakhir per satker
        df_std = df_std.sort_values(["Kode Satker", "Tanggal Posting Revisi"], ascending=[True, False])
        df_std = df_std.drop_duplicates(subset="Kode Satker", keep="first")

        # 9️⃣ Simpan ke session_state
        if "DATA_DIPA_by_year" not in st.session_state:
            st.session_state.DATA_DIPA_by_year = {}

        st.session_state.DATA_DIPA_by_year[int(tahun_dipa)] = df_std.copy()

        # 🔟 Upload ke GitHub
        with st.spinner("Mengunggah ke GitHub..."):
            out = io.BytesIO()
            with pd.ExcelWriter(out, engine="openpyxl") as writer:
                df_std.to_excel(writer, index=False, sheet_name=f"DIPA_{tahun_dipa}")

                # Header styling
                ws = writer.sheets[f"DIPA_{tahun_dipa}"]
                for cell in ws[1]:
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="366092", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center", vertical="center")

            out.seek(0)
            save_file_to_github(out.getvalue(), f"DIPA_{tahun_dipa}.xlsx", "DATA_DIPA")

        # Preview
        st.write("**Preview 5 baris pertama:**")
        st.dataframe(df_std.head(5))

        return df_std, int(tahun_dipa), "✅ Sukses diproses"

    except Exception as e:
        st.error(f"❌ Error: {str(e)}")
        import traceback
        st.code(traceback.format_exc())
        return None, None, f"❌ Error: {str(e)}"

    
import streamlit as st
import pandas as pd
import io
from github import Github, Auth
import base64

# Fungsi bantu
def get_latest_dipa(dipa_df):
    if 'Tanggal Posting Revisi' in dipa_df.columns:
        dipa_df['Tanggal Posting Revisi'] = pd.to_datetime(dipa_df['Tanggal Posting Revisi'], errors='coerce')
        latest_dipa = dipa_df.sort_values('Tanggal Posting Revisi', ascending=False) \
                              .drop_duplicates(subset='Kode Satker', keep='first')
    elif 'No Revisi Terakhir' in dipa_df.columns:
        latest_dipa = dipa_df.sort_values('No Revisi Terakhir', ascending=False) \
                              .drop_duplicates(subset='Kode Satker', keep='first')
    else:
        latest_dipa = dipa_df.drop_duplicates(subset='Kode Satker', keep='first')
    return latest_dipa

def merge_ikpa_dipa_auto():
    
    if st.session_state.get("ikpa_dipa_merged", False):
        return

    if "data_storage" not in st.session_state:
        return

    if "DATA_DIPA_by_year" not in st.session_state:
        return

    for (bulan, tahun), df_ikpa in st.session_state.data_storage.items():

        dipa = st.session_state.DATA_DIPA_by_year.get(int(tahun))
        if dipa is None:
            continue

        df_final = df_ikpa.copy()
        dipa_latest = get_latest_dipa(dipa)

        # NORMALISASI KUNCI (WAJIB)
        df_final["Kode Satker"] = df_final["Kode Satker"].astype(str).str.zfill(6)
        dipa_latest["Kode Satker"] = dipa_latest["Kode Satker"].astype(str).str.zfill(6)

        dipa_selected = dipa_latest[['Kode Satker', 'Total Pagu', 'Jenis Satker']]

        df_final = df_final.drop(columns=['Total Pagu', 'Jenis Satker'], errors='ignore')

        df_merged = pd.merge(
            df_final,
            dipa_selected,
            on='Kode Satker',
            how='left'
        )

        st.session_state.data_storage[(bulan, tahun)] = df_merged

    st.session_state.ikpa_dipa_merged = True


# ============================================================
# 🔹 Fungsi convert DataFrame ke Excel bytes
# ============================================================
def to_excel_bytes(df):
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="xlsxwriter") as writer:
        df.to_excel(writer, index=False)
    return output.getvalue()

# ============================================================
# 🔹 Fungsi push file ke GitHub
# ============================================================
def push_to_github(file_bytes, repo_path, repo_name, token, commit_message):
    g = Github(auth=Auth.Token(token))
    repo = g.get_repo(repo_name)

    try:
        # Cek apakah file sudah ada
        try:
            contents = repo.get_contents(repo_path)
            repo.update_file(contents.path, commit_message, file_bytes, contents.sha)
            st.success(f"✅ File {repo_path} berhasil diupdate di GitHub")
        except Exception as e_inner:
            # Jika file belum ada atau path salah, buat baru
            repo.create_file(repo_path, commit_message, file_bytes)
            st.success(f"✅ File {repo_path} berhasil dibuat di GitHub")
    except Exception as e:
        st.error(f"❌ Gagal push ke GitHub: {e}")
        
# Deteksi IKPA KPPN
def detect_header_row(excel_file, keyword, max_rows=10):
    """
    Mendeteksi baris header berdasarkan keyword kolom
    """
    preview = pd.read_excel(excel_file, header=None, nrows=max_rows)

    for i in range(len(preview)):
        row = preview.iloc[i].astype(str).str.strip()
        if keyword in row.values:
            return i
    return None

# ============================================================
#  Menu Admin
# ============================================================
def page_admin():
    st.title("🔐 Halaman Administrasi")

    # ===============================
    # 🔑 LOGIN ADMIN
    # ===============================
    if "authenticated" not in st.session_state:
        st.session_state.authenticated = False

    if not st.session_state.authenticated:
        st.warning("🔒 Halaman ini memerlukan autentikasi Admin")
        password = st.text_input("Masukkan Password Admin", type="password")

        if st.button("Login"):
            if password == ADMIN_PASSWORD:
                st.session_state.authenticated = True
                st.success("✔ Login berhasil")
                st.rerun()
            else:
                st.error("❌ Password salah")
        return

    st.success("✔ Anda login sebagai Admin")

    # ===============================
    # 🔄 KONTROL DATA (MANUAL OVERRIDE)
    # ===============================
    st.subheader("Manajemen Data")

    # ============================================================
    # JIKA DATA SUDAH SIAP (MERGE BERHASIL)
    # ============================================================
    if st.session_state.get("ikpa_dipa_merged", False):

        st.success(" Data IKPA & DIPA sudah siap digunakan dan merge berhasil")
        st.caption("Tidak diperlukan proses atau tindakan Admin")

    # ============================================================
    #  JIKA DATA BELUM SIAP (BELUM MERGE / GAGAL)
    # ============================================================
    else:

        st.warning("⚠️ Data belum siap atau perlu diproses")

        # Tombol proses awal
        if st.button("🔄 Load & Olah Data"):
            with st.spinner(" Memuat & menggabungkan data..."):
                st.session_state.ikpa_dipa_merged = False
                load_DATA_DIPA_from_github()
                load_data_from_github()
                merge_ikpa_dipa_auto()
            st.success("✅ Proses selesai")
            st.rerun()

        # Reset hanya muncul kalau data ada tapi merge gagal
        if st.session_state.get("data_storage") or st.session_state.get("DATA_DIPA_by_year"):
            with st.expander(" Admin Lanjutan (Opsional)"):
                if st.button(" Reset Status Merge"):
                    st.session_state.ikpa_dipa_merged = False
                    st.warning(" Status merge direset. Data akan diproses ulang.")
                    st.rerun()


    # ===============================
    # 🔍 SIDEBAR DEBUG
    # ===============================
    with st.sidebar:
        st.markdown("### 🔍 Debug DIPA")
        if st.button("Cek Status DIPA"):
            if "DATA_DIPA_by_year" in st.session_state:
                for tahun, df in st.session_state.DATA_DIPA_by_year.items():
                    st.write(f"**{tahun}:** {len(df)} baris")
                    st.write(f"- Kode Satker kosong: {df['Kode Satker'].eq('').sum()}")
                    st.write(f"- Total Pagu = 0: {df['Total Pagu'].eq(0).sum()}")
            else:
                st.warning("DATA_DIPA_by_year belum dimuat")

    # ===============================
    # 📌 TAB MENU
    # ===============================
    tab1, tab2, tab3, tab4, tab5 = st.tabs([
        "📤 Upload Data",
        "🗑️ Hapus Data",
        "📥 Download Data",
        "📋 Download Template",
        "🕓 Riwayat Aktivitas"
    ])

    # ============================================================
    # TAB 1: UPLOAD DATA (IKPA, DIPA, Referensi)
    # ============================================================
    with tab1:
        # Upload Data IKPA Satker
        st.subheader("📤 Upload Data IKPA Satker")

        upload_year = st.selectbox(
            "Pilih Tahun",
            list(range(2020, 2031)),
            index=list(range(2020, 2031)).index(datetime.now().year)
        )

        uploaded_files = st.file_uploader(
            "Pilih satu atau beberapa file Excel IKPA Satker",
            type=["xlsx", "xls"],
            accept_multiple_files=True
        )

        if uploaded_files:

            st.info("📄 File yang diupload:")
            for f in uploaded_files:
                st.write("•", f.name)

            if st.button("🔄 Proses Semua Data IKPA", type="primary"):

                with st.spinner("Memproses semua file IKPA Satker..."):

                    for uploaded_file in uploaded_files:
                        try:
                            # ======================
                            # 🔄 PROSES FILE (PARSER LENGKAP)
                            # ======================
                            uploaded_file.seek(0)
                            df_final, month, year = process_excel_file(
                                uploaded_file,
                                upload_year
                            )

                            if df_final is None or month == "UNKNOWN":
                                st.warning(
                                    f"⚠️ {uploaded_file.name} gagal diproses "
                                    f"(bulan tidak terdeteksi)"
                                )
                                continue

                            # ======================
                            # NORMALISASI KODE SATKER
                            # ======================
                            if "Kode Satker" in df_final.columns:
                                df_final["Kode Satker"] = (
                                    df_final["Kode Satker"]
                                    .astype(str)
                                    .apply(normalize_kode_satker)
                                )

                            # ======================
                            # OVERRIDE JIKA BULAN SAMA
                            # ======================
                            st.session_state.data_storage.pop(
                                (month, str(year)), None
                            )

                            # ======================
                            # REGISTRASI KE SISTEM (KUNCI)
                            # ======================
                            register_ikpa_satker(
                                df_final,
                                month,
                                year,
                                source="Manual"
                            )

                            # tandai perlu merge ulang
                            need_merge = True
                            st.session_state.ikpa_dipa_merged = False

                            # ======================
                            # 💾 SIMPAN KE GITHUB
                            # ======================
                            excel_bytes = io.BytesIO()
                            with pd.ExcelWriter(
                                excel_bytes,
                                engine="openpyxl"
                            ) as writer:
                                df_final.to_excel(
                                    writer,
                                    index=False,
                                    sheet_name="Data IKPA"
                                )
                            excel_bytes.seek(0)

                            save_file_to_github(
                                excel_bytes.getvalue(),
                                f"IKPA_{month}_{year}.xlsx",
                                folder="data"
                            )

                            st.session_state.activity_log.append({
                                "Waktu": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                                "Aksi": "Upload IKPA Satker",
                                "Periode": f"{month} {year}",
                                "Status": "✅ Sukses"
                            })

                            st.success(
                                f"✅ {uploaded_file.name} → "
                                f"{month} {year} berhasil diproses"
                            )

                        except Exception as e:
                            st.error(f"❌ Error {uploaded_file.name}: {e}")

                    if need_merge and st.session_state.DATA_DIPA_by_year and not st.session_state.get("ikpa_dipa_merged", False):
                        with st.spinner("🔄 Menggabungkan IKPA & DIPA..."):
                            merge_ikpa_dipa_auto()
                            st.session_state.ikpa_dipa_merged = True

        
        # Submenu Upload Data IKPA KPPN
        st.subheader("📝 Upload Data IKPA KPPN")
        # ===============================
        # 📅 PILIH TAHUN
        # ===============================
        upload_year_kppn = st.selectbox(
            "Pilih Tahun",
            list(range(2020, 2031)),
            index=list(range(2020, 2031)).index(datetime.now().year),
            key="tahun_kppn"
        )

        # ===============================
        # 📂 UPLOAD FILE
        # ===============================
        uploaded_file_kppn = st.file_uploader(
            "Pilih file Excel IKPA KPPN",
            type=["xlsx", "xls"],
            key="file_kppn"
        )

        # ===============================
        # 🔐 INISIALISASI SESSION
        # ===============================
        if "data_storage_kppn" not in st.session_state:
            st.session_state.data_storage_kppn = {}

        # ===============================
        # 🚦 VALIDASI FILE
        # ===============================
        if uploaded_file_kppn is not None:
            try:
                # Cari baris header otomatis
                header_row = find_header_row_by_keyword(
                    uploaded_file_kppn,
                    keyword="Nama KPPN"
                )

                if header_row is None:
                    st.error(
                        "GAGAL UPLOAD!\n\n"
                        "Kolom **'Nama KPPN'** tidak ditemukan.\n"
                        "File ini BUKAN data IKPA KPPN yang valid."
                    )
                    st.stop()

                # Baca data dengan header yang benar
                uploaded_file_kppn.seek(0)
                df_check = pd.read_excel(
                    uploaded_file_kppn,
                    header=header_row
                )

                # Normalisasi nama kolom
                df_check.columns = (
                    df_check.columns.astype(str)
                    .str.strip()
                    .str.replace(r"\s+", " ", regex=True)
                )

                #  SALAH FILE: IKPA SATKER
                if "Nama Satker" in df_check.columns:
                    st.error(
                        "GAGAL UPLOAD!\n\n"
                        "File yang Anda upload adalah **IKPA SATKER**.\n"
                        "Halaman ini hanya menerima **IKPA KPPN**."
                    )
                    st.stop()

                # ===============================
                # 🔍 DETEKSI BULAN (HEADER ATAS)
                # ===============================
                uploaded_file_kppn.seek(0)
                df_info = pd.read_excel(uploaded_file_kppn, header=None)

                month_preview = "UNKNOWN"
                MONTH_MAP = {
                    "JAN": "JANUARI", "JANUARI": "JANUARI",
                    "FEB": "FEBRUARI", "FEBRUARI": "FEBRUARI",
                    "MAR": "MARET", "MARET": "MARET",
                    "APR": "APRIL", "APRIL": "APRIL",
                    "MEI": "MEI",
                    "JUN": "JUNI", "JUNI": "JUNI",
                    "JUL": "JULI", "JULI": "JULI",
                    "AGT": "AGUSTUS", "AGS": "AGUSTUS", "AGUSTUS": "AGUSTUS",
                    "SEP": "SEPTEMBER", "SEPTEMBER": "SEPTEMBER",
                    "OKT": "OKTOBER", "OKTOBER": "OKTOBER",
                    "NOV": "NOVEMBER", "NOVEMBER": "NOVEMBER",
                    "DES": "DESEMBER", "DESEMBER": "DESEMBER"
                }

                if df_info.shape[0] > 1:
                    text = str(df_info.iloc[1, 0]).upper()
                    for k, v in MONTH_MAP.items():
                        if k in text:
                            month_preview = v
                            break

                period_key_preview = (month_preview, str(upload_year_kppn))

                if period_key_preview in st.session_state.data_storage_kppn:
                    st.warning(
                        f" Data IKPA KPPN **{month_preview} {upload_year_kppn}** sudah ada."
                    )
                    confirm_replace = st.checkbox(
                        " Ganti data yang sudah ada",
                        key=f"confirm_replace_kppn_{month_preview}_{upload_year_kppn}"
                    )
                else:
                    confirm_replace = True
                    st.info(
                        f"Akan mengunggah data IKPA KPPN "
                        f"untuk periode **{month_preview} {upload_year_kppn}**"
                    )

            except Exception as e:
                st.error(f" Gagal membaca file: {e}")
                confirm_replace = False

            # ===============================
            # 🔄 PROSES DATA
            # ===============================
            if st.button(
                " Proses Data IKPA KPPN",
                type="primary",
                disabled=not confirm_replace,
                key="proses_kppn"
            ):
                with st.spinner("Memproses data IKPA KPPN..."):

                    df_processed, month, year = process_excel_file_kppn(
                        uploaded_file_kppn,
                        upload_year_kppn
                    )

                    if df_processed is None:
                        st.error(" Gagal memproses file IKPA KPPN.")
                        st.stop()

                    period_key = (str(month), str(year))
                    filename = f"IKPA_KPPN_{month}_{year}.xlsx"

                    try:
                        #  Simpan ke session
                        st.session_state.data_storage_kppn[period_key] = df_processed

                        #  Simpan ke GitHub
                        excel_bytes = io.BytesIO()
                        with pd.ExcelWriter(excel_bytes, engine="openpyxl") as writer:
                            df_processed.drop(
                                ["Bobot", "Nilai Terbobot"],
                                axis=1,
                                errors="ignore"
                            ).to_excel(
                                writer,
                                index=False,
                                sheet_name="Data IKPA KPPN"
                            )
                        excel_bytes.seek(0)

                        save_file_to_github(
                            excel_bytes.getvalue(),
                            filename,
                            folder="data_kppn"
                        )

                        st.success(
                            f" Data IKPA KPPN {month} {year} berhasil disimpan."
                        )
                        st.snow()

                    except Exception as e:
                        st.error(f" Gagal menyimpan ke GitHub: {e}")
            
        # ============================================================
        # SUBMENU: UPLOAD DATA DIPA
        # ============================================================
        st.markdown("---")
        st.subheader("📤 Upload Data DIPA")

        uploaded_dipa_file = st.file_uploader(
            "Pilih file Excel DIPA (mentah dari SAS/SMART/Kemenkeu)",
            type=['xlsx', 'xls'],
            key="upload_dipa"
        )

        # Tombol proses DIPA
        if uploaded_dipa_file is not None:
            if st.button("🔄 Proses Data DIPA", type="primary"):
                with st.spinner("Memproses data DIPA..."):

                    try:
                        # 1️⃣ Proses file raw DIPA → dibersihkan → revisi terbaru
                        df_clean, tahun_dipa, status_msg = process_uploaded_dipa(uploaded_dipa_file, save_file_to_github)

                        if df_clean is None:
                            st.error(f"❌ Gagal memproses DIPA: {status_msg}")
                            st.stop()

                        # 2️⃣ Pastikan kolom Kode Satker distandardkan
                        df_clean["Kode Satker"] = df_clean["Kode Satker"].astype(str).apply(normalize_kode_satker)

                        # 3️⃣ Simpan ke session_state per tahun
                        if "DATA_DIPA_by_year" not in st.session_state:
                            st.session_state.DATA_DIPA_by_year = {}

                        st.session_state.DATA_DIPA_by_year[int(tahun_dipa)] = df_clean.copy()

                        # 4️⃣ Simpan ke GitHub dalam folder `DATA_DIPA`
                        excel_bytes = io.BytesIO()
                        with pd.ExcelWriter(excel_bytes, engine='openpyxl') as writer:
                            df_clean.to_excel(writer, index=False, sheet_name=f"DIPA_{tahun_dipa}")

                        excel_bytes.seek(0)

                        save_file_to_github(
                            excel_bytes.getvalue(),
                            f"DIPA_{tahun_dipa}.xlsx",  
                            folder="DATA_DIPA"
                        )

                        # 5️⃣ Catat log
                        st.session_state.activity_log.append({
                            "Waktu": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                            "Aksi": "Upload DIPA",
                            "Periode": f"Tahun {tahun_dipa}",
                            "Status": "Sukses"
                        })

                        # 6️⃣ Tampilkan hasil preview
                        st.success(f"✅ Data DIPA tahun {tahun_dipa} berhasil diproses & disimpan.")
                        st.dataframe(df_clean.head(10), use_container_width=True)

                    except Exception as e:
                        st.error(f"❌ Terjadi error saat memproses file DIPA: {e}")

        # ============================================================
        # SUBMENU: Upload Data Referensi
        # ============================================================
        st.markdown("---")
        st.subheader("📚 Upload / Perbarui Data Referensi Satker & K/L")
        st.info("""
        - File referensi ini berisi kolom: **Kode BA, K/L, Kode Satker, Uraian Satker-SINGKAT, Uraian Satker-LENGKAP**  
        - Saat diupload, sistem akan **menggabungkan** dengan data lama:  
        🔹 Jika `Kode Satker` sudah ada → baris lama akan **diganti**  
        🔹 Jika `Kode Satker` belum ada → akan **ditambahkan baru**
        """)

        uploaded_ref = st.file_uploader(
            "📤 Pilih File Data Referensi Satker & K/L",
            type=['xlsx', 'xls'],
            key="ref_upload"
        )

        if uploaded_ref is not None:
            try:
                new_ref = pd.read_excel(uploaded_ref)
                new_ref.columns = [c.strip() for c in new_ref.columns]

                required = ['Kode BA', 'K/L', 'Kode Satker', 'Uraian Satker-SINGKAT', 'Uraian Satker-LENGKAP']
                if not all(col in new_ref.columns for col in required):
                    st.error("❌ Kolom wajib tidak lengkap dalam file referensi.")
                    st.stop()

                new_ref['Kode Satker'] = new_ref['Kode Satker'].apply(normalize_kode_satker)

                # Gabungkan atau buat baru
                if 'reference_df' in st.session_state:
                    old_ref = st.session_state.reference_df.copy()

                    # 🔹 Normalize old reference too (critical!)
                    if 'Kode Satker' in old_ref.columns:
                        old_ref['Kode Satker'] = old_ref['Kode Satker'].apply(normalize_kode_satker)

                    # 🔹 Combine and deduplicate
                    merged = pd.concat([old_ref, new_ref], ignore_index=True)
                    merged = merged.drop_duplicates(subset=['Kode Satker'], keep='last')

                    # 🔹 Optional: enforce consistent string stripping
                    merged['Kode Satker'] = merged['Kode Satker'].astype(str).str.strip()

                    st.session_state.reference_df = merged
                    st.success(f"✅ Data Referensi diperbarui ({len(merged)} total baris).")
                else:
                    st.session_state.reference_df = new_ref
                    st.success(f"✅ Data Referensi baru dimuat ({len(new_ref)} baris).")

                st.dataframe(st.session_state.reference_df.tail(10), use_container_width=True)

                # Save merged reference data permanently to GitHub
                try:
                    excel_bytes_ref = io.BytesIO()
                    with pd.ExcelWriter(excel_bytes_ref, engine='openpyxl') as writer:
                        st.session_state.reference_df.to_excel(
                            writer, index=False, sheet_name='Data Referensi',
                            startrow=0, startcol=0  # ✅ PERBAIKAN
                        )
                        
                        # Format header
                        workbook = writer.book
                        worksheet = writer.sheets['Data Referensi']
                        
                        for cell in worksheet[1]:
                            cell.font = Font(bold=True, color="FFFFFF")
                            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                            cell.alignment = Alignment(horizontal="center", vertical="center")
                    
                    excel_bytes_ref.seek(0)

                    save_file_to_github(
                        excel_bytes_ref.getvalue(),
                        "Template_Data_Referensi.xlsx",
                        folder="templates"
                    )
                    st.success("💾 Data Referensi berhasil disimpan ke GitHub (templates/Template_Data_Referensi.xlsx).")
                except Exception as e:
                    st.error(f"❌ Gagal menyimpan Data Referensi ke GitHub: {e}")

            except Exception as e:
                st.error(f"❌ Gagal memproses Data Referensi: {e}")

    # ============================================================
    # TAB 2: HAPUS DATA
    # ============================================================
    with tab2:
        # Submenu Hapus Data IKPA Satker
        st.subheader("🗑️ Hapus Data IKPA Satker")
        if not st.session_state.data_storage:
            st.info("ℹ️ Belum ada data IKPA tersimpan.")
        else:
            available_periods = sorted(st.session_state.data_storage.keys(), reverse=True)
            period_to_delete = st.selectbox(
                "Pilih periode yang akan dihapus",
                options=available_periods,
                format_func=lambda x: f"{x[0].capitalize()} {x[1]}"
            )
            month, year = period_to_delete
            filename = f"data/IKPA_{month}_{year}.xlsx"

            confirm_delete = st.checkbox(
                f"⚠️ Hapus data {month} {year} dari sistem dan GitHub.",
                key=f"confirm_delete_{month}_{year}"
            )

            if st.button("🗑️ Hapus Data IKPA Satker", type="primary") and confirm_delete:
                try:
                    del st.session_state.data_storage[period_to_delete]
                    token = st.secrets.get("GITHUB_TOKEN")
                    repo_name = st.secrets.get("GITHUB_REPO")
                    g = Github(auth=Auth.Token(token))
                    repo = g.get_repo(repo_name)
                    contents = repo.get_contents(f"data/IKPA_{month}_{year}.xlsx")
                    repo.delete_file(contents.path, f"Delete {filename}", contents.sha)
                    st.success(f"✅ Data {month} {year} dihapus dari sistem & GitHub.")
                    st.snow()
                    st.session_state.activity_log.append({
                        "Waktu": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                        "Aksi": "Hapus IKPA",
                        "Periode": f"{month} {year}",
                        "Status": "✅ Sukses"
                    })
                except Exception as e:
                    st.error(f"❌ Gagal menghapus data: {e}")
                    
        # Submenu Hapus Data IKPA KPPN
        st.subheader("🗑️ Hapus Data IKPA KPPN")

        try:
            token = st.secrets["GITHUB_TOKEN"]
            repo_name = st.secrets["GITHUB_REPO"]

            g = Github(auth=Auth.Token(token))
            repo = g.get_repo(repo_name)

            # Ambil semua file di folder data_kppn
            contents = repo.get_contents("data_kppn")

            files_kppn = [
                c.name for c in contents
                if c.name.startswith("IKPA_KPPN_") and c.name.endswith(".xlsx")
            ]

        except Exception as e:
            st.error(f"❌ Gagal membaca data dari GitHub: {e}")
            st.stop()

        # ===============================
        # JIKA BELUM ADA DATA
        # ===============================
        if not files_kppn:
            st.info("ℹ️ Belum ada data IKPA KPPN tersimpan.")
            st.stop()

        # ===============================
        # PILIH FILE
        # ===============================
        selected_file = st.selectbox(
            "Pilih data IKPA KPPN yang akan dihapus",
            sorted(files_kppn, reverse=True)
        )

        confirm_delete = st.checkbox(
            f"⚠️ Saya yakin ingin menghapus **{selected_file}** dari sistem dan GitHub"
        )

        # ===============================
        # PROSES HAPUS
        # ===============================
        if st.button("🗑️ Hapus Data IKPA KPPN", type="primary") and confirm_delete:
            try:
                file_path = f"data_kppn/{selected_file}"
                content = repo.get_contents(file_path)

                repo.delete_file(
                    content.path,
                    f"Delete {selected_file}",
                    content.sha
                )

                # Log aktivitas
                if "activity_log" not in st.session_state:
                    st.session_state.activity_log = []

                st.session_state.activity_log.append({
                    "Waktu": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    "Aksi": "Hapus IKPA KPPN",
                    "File": selected_file,
                    "Status": "✅ Sukses"
                })

                st.success(f"✅ {selected_file} berhasil dihapus.")
                st.snow()
                st.rerun()

            except Exception as e:
                st.error(f"❌ Gagal menghapus data IKPA KPPN: {e}")


        # Submenu Hapus Data DIPA
        st.markdown("---")
        st.subheader("🗑️ Hapus Data DIPA")
        if not st.session_state.get("DATA_DIPA_by_year"):
            st.info("ℹ️ Belum ada data DIPA tersimpan.")
        else:
            available_years = sorted(st.session_state.DATA_DIPA_by_year.keys(), reverse=True)
            year_to_delete = st.selectbox(
                "Pilih tahun DIPA yang akan dihapus",
                options=available_years,
                format_func=lambda x: f"Tahun {x}",
                key="delete_dipa_year"
            )
            filename_dipa = f"DATA_DIPA/DIPA_{year_to_delete}.xlsx"

            confirm_delete_dipa = st.checkbox(
                f"⚠️ Hapus data DIPA tahun {year_to_delete} dari sistem dan GitHub.",
                key=f"confirm_delete_dipa_{year_to_delete}"
            )

            if st.button("🗑️ Hapus Data DIPA Ini", type="primary", key="btn_delete_dipa") and confirm_delete_dipa:
                try:
                    del st.session_state.DATA_DIPA_by_year[year_to_delete]
                    token = st.secrets.get("GITHUB_TOKEN")
                    repo_name = st.secrets.get("GITHUB_REPO")
                    g = Github(auth=Auth.Token(token))
                    repo = g.get_repo(repo_name)
                    contents = repo.get_contents(filename_dipa)
                    repo.delete_file(contents.path, f"Delete {filename_dipa}", contents.sha)
                    st.success(f"✅ Data DIPA tahun {year_to_delete} dihapus dari sistem & GitHub.")
                    st.snow()
                    st.session_state.activity_log.append({
                        "Waktu": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                        "Aksi": "Hapus DIPA",
                        "Periode": f"Tahun {year_to_delete}",
                        "Status": "✅ Sukses"
                    })
                except Exception as e:
                    st.error(f"❌ Gagal menghapus data DIPA: {e}")

    # ============================================================
    # TAB 3: DOWNLOAD DATA
    # ============================================================
    with tab3:
        st.subheader("📥 Download IKPA Satker")

        if "data_storage" not in st.session_state or not st.session_state.data_storage:
            st.info("🔹 Data belum tersedia untuk diunduh")
        else:
            available_periods = sorted(st.session_state.data_storage.keys(), reverse=True)
            period_to_download = st.selectbox(
                "Pilih periode untuk download",
                options=available_periods,
                format_func=lambda x: f"{x[0]} {x[1]}"
            )

            df_selected = st.session_state.data_storage.get(period_to_download)
            if df_selected is not None:
                filename = f"IKPA_{period_to_download[0]}_{period_to_download[1]}.xlsx"
                excel_bytes = to_excel_bytes(df_selected)  # pastikan fungsi ini sudah ada
                st.download_button(
                    label=f"Download {filename}",
                    data=excel_bytes,
                    file_name=filename,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
         
        # ===========================
        # Submenu Download Data IKPA KPPN
        # ===========================
        st.subheader("📥 Download Data IKPA KPPN")

        try:
            token = st.secrets["GITHUB_TOKEN"]
            repo_name = st.secrets["GITHUB_REPO"]

            g = Github(auth=Auth.Token(token))
            repo = g.get_repo(repo_name)

            contents = repo.get_contents("data_kppn")

            files_kppn = [
                c.name for c in contents
                if c.name.startswith("IKPA_KPPN_") and c.name.endswith(".xlsx")
            ]

        except Exception as e:
            st.error(f"❌ Gagal membaca data dari GitHub: {e}")
            st.stop()

        # ===============================
        # JIKA BELUM ADA DATA
        # ===============================
        if not files_kppn:
            st.info("ℹ️ Belum ada data IKPA KPPN tersedia untuk diunduh.")
            st.stop()

        # ===============================
        # PILIH FILE
        # ===============================
        selected_file = st.selectbox(
            "Pilih data IKPA KPPN",
            sorted(files_kppn, reverse=True)
        )

        # ===============================
        # AMBIL FILE DARI GITHUB
        # ===============================
        try:
            file_path = f"data_kppn/{selected_file}"
            file_content = repo.get_contents(file_path)
            file_bytes = file_content.decoded_content
        except Exception as e:
            st.error(f"❌ Gagal mengambil file: {e}")
            st.stop()

        # ===============================
        # DOWNLOAD BUTTON
        # ===============================
        st.download_button(
            label="📥 Download File IKPA KPPN",
            data=file_bytes,
            file_name=selected_file,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)
                
        # ===========================
        # Submenu Download Data DIPA
        # ===========================

        st.markdown("### 📥 Download Data DIPA")

        if not st.session_state.get("DATA_DIPA_by_year"):
            st.info("ℹ️ Belum ada data DIPA.")
        else:
            available_years = sorted(st.session_state.DATA_DIPA_by_year.keys(), reverse=True)

            year_to_download = st.selectbox(
                "Pilih tahun DIPA",
                options=available_years,
                format_func=lambda x: f"Tahun {x}",
                key="download_dipa_year"
            )

            # Ambil data yang sudah bersih dari load()
            df = st.session_state.DATA_DIPA_by_year[year_to_download].copy()

            # Kolom yang ingin ditampilkan
            desired_columns = [
                "Kode Satker",
                "Satker",
                "Tahun",
                "Tanggal Posting Revisi",
                "Total Pagu",
                "Jenis Satker",
                "NO",
                "Kementerian",
                "Kode Status History",
                "Jenis Revisi",
                "Revisi ke-",
                "No Dipa",
                "Tanggal Dipa",
                "Owner",
                "Digital Stamp"
            ]

            # Filter kolom yang ada
            df = df[[c for c in desired_columns if c in df.columns]]

            # Ambil revisi terbaru
            if "Kode Satker" in df.columns and "Tanggal Posting Revisi" in df.columns:
                df["Tanggal Posting Revisi"] = pd.to_datetime(df["Tanggal Posting Revisi"], errors="coerce")
                df = df.sort_values(
                    by=["Kode Satker", "Tanggal Posting Revisi"],
                    ascending=[True, False]
                ).drop_duplicates(subset="Kode Satker", keep="first")

            # Klasifikasi Satker
            if "Total Pagu" in df.columns:
                q40 = df["Total Pagu"].quantile(0.40)
                q70 = df["Total Pagu"].quantile(0.70)

                def klasifikasi(x):
                    if x >= q70: return "Satker Besar"
                    elif x >= q40: return "Satker Sedang"
                    return "Satker Kecil"

                df["Jenis Satker"] = df["Total Pagu"].apply(klasifikasi)

            # Preview
            with st.expander("Preview Data"):
                st.dataframe(df.head(10), use_container_width=True)

            # Export Excel
            output = io.BytesIO()
            with pd.ExcelWriter(output, engine="openpyxl") as writer:
                df.to_excel(writer, index=False, sheet_name=f"DIPA_{year_to_download}")

            output.seek(0)

            st.download_button(
                "📥 Download Excel DIPA",
                data=output,
                file_name=f"DIPA_{year_to_download}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )



        # Download Data Satker Tidak Terdaftar
        st.markdown("---")
        st.subheader("📥 Download Data Satker yang Belum Terdaftar di Tabel Referensi")
        
        if st.button("📥 Generate & Download Laporan"):
            st.info("ℹ️ Fitur ini menggunakan data dari session state untuk performa optimal.")

    # ============================================================
    # TAB 4: DOWNLOAD TEMPLATE
    # ============================================================
    with tab4:
        st.subheader("📋 Download Template")
        st.markdown("### 📘 Template IKPA")
        try:
            token = st.secrets["GITHUB_TOKEN"]
            repo_name = st.secrets["GITHUB_REPO"]
            g = Github(auth=Auth.Token(token))
            repo = g.get_repo(repo_name)
            file_content = repo.get_contents("templates/Template_IKPA.xlsx")
            template_data = base64.b64decode(file_content.content)
        except Exception:
            template_data = get_template_file()

        if template_data:
            st.download_button(
                label="📥 Download Template IKPA",
                data=template_data,
                file_name="Template_IKPA.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

        st.markdown("---")
        st.markdown("### 📗 Template Data Referensi Satker & K/L")

        # 🧩 Use latest reference data for template content
        if 'reference_df' in st.session_state and not st.session_state.reference_df.empty:
            template_ref = st.session_state.reference_df.copy()
        else:
            # fallback: try load from GitHub
            try:
                token = st.secrets["GITHUB_TOKEN"]
                repo_name = st.secrets["GITHUB_REPO"]
                g = Github(auth=Auth.Token(token))
                repo = g.get_repo(repo_name)
                ref_content = repo.get_contents("templates/Template_Data_Referensi.xlsx")
                ref_data = base64.b64decode(ref_content.content)
                template_ref = pd.read_excel(io.BytesIO(ref_data))
            except Exception:
                template_ref = pd.DataFrame({
                    'No': [],
                    'Kode BA': [],
                    'K/L': [],
                    'Kode Satker': [],
                    'Uraian Satker-SINGKAT': [],
                    'Uraian Satker-LENGKAP': []
                })

        output_ref = io.BytesIO()
        with pd.ExcelWriter(output_ref, engine='openpyxl') as writer:
            # ✅ PERBAIKAN: Mulai dari A1
            template_ref.to_excel(writer, index=False, sheet_name='Data Referensi',
                                  startrow=0, startcol=0)
            
            # Format header
            workbook = writer.book
            worksheet = writer.sheets['Data Referensi']
            
            for cell in worksheet[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
        
        output_ref.seek(0)

        st.download_button(
            label="📥 Download Template Data Referensi",
            data=output_ref,
            file_name="Template_Data_Referensi.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

# ===============================
# MAIN APP
# ===============================
def main():

    # ============================================================
    # 1️⃣ LOAD REFERENCE DATA (SEKALI SAJA)
    # ============================================================
    if "reference_df" not in st.session_state:

        token = st.secrets.get("GITHUB_TOKEN")
        repo_name = st.secrets.get("GITHUB_REPO")

        if not token or not repo_name:
            st.session_state.reference_df = pd.DataFrame({
                'Kode BA': [], 'K/L': [], 'Kode Satker': [],
                'Uraian Satker-SINGKAT': [], 'Uraian Satker-LENGKAP': []
            })
        else:
            try:
                g = Github(auth=Auth.Token(token))
                repo = g.get_repo(repo_name)
                ref_path = "templates/Template_Data_Referensi.xlsx"

                ref_file = repo.get_contents(ref_path)
                ref_data = base64.b64decode(ref_file.content)

                ref_df = pd.read_excel(io.BytesIO(ref_data))
                ref_df.columns = [c.strip() for c in ref_df.columns]

                st.session_state.reference_df = ref_df

            except Exception:
                st.session_state.reference_df = pd.DataFrame({
                    'Kode BA': [], 'K/L': [], 'Kode Satker': [],
                    'Uraian Satker-SINGKAT': [], 'Uraian Satker-LENGKAP': []
                })

    # ============================================================
    # 2️⃣ AUTO LOAD DATA IKPA
    # ============================================================
    if not st.session_state.data_storage:
        with st.spinner("🔄 Memuat data IKPA..."):
            load_data_from_github()

    # ============================================================
    # 3️⃣ AUTO LOAD DATA DIPA (HASIL PROCESSING STREAMLIT)
    # ============================================================
    if not st.session_state.DATA_DIPA_by_year:
        with st.spinner("🔄 Memuat data DIPA..."):
            load_DATA_DIPA_from_github()

    # ============================================================
    # 4️⃣ FINALISASI DATA DIPA (AMAN)
    # ============================================================
    if st.session_state.DATA_DIPA_by_year:
        for tahun, df in st.session_state.DATA_DIPA_by_year.items():
            df = df.copy()
            if "Uraian Satker" in df.columns:
                df["Uraian Satker-RINGKAS"] = (
                    df["Uraian Satker"]
                    .fillna("-")
                    .astype(str)
                    .str[:30]
                )
            else:
                df["Uraian Satker-RINGKAS"] = "-"
            st.session_state.DATA_DIPA_by_year[tahun] = df

    # ============================================================
    # 5️⃣ AUTO MERGE IKPA + DIPA 
    # ============================================================
    if (
        st.session_state.data_storage and
        st.session_state.DATA_DIPA_by_year and
        not st.session_state.ikpa_dipa_merged
    ):
        with st.spinner("🔄 Menggabungkan data IKPA & DIPA..."):
            merge_ikpa_dipa_auto()
            
    # ============================================================
    # NOTIF GLOBAL STATUS DATA (MUNCUL SAAT APP DIBUKA)
    # ============================================================
    if st.session_state.get("ikpa_dipa_merged", False):
        st.success(" Data IKPA & DIPA berhasil dimuat dan siap digunakan")


    # ============================================================
    # Sidebar + Routing halaman
    # ============================================================
    st.sidebar.title("🧭 Navigasi")
    st.sidebar.markdown("---")

    if "page" not in st.session_state:
        st.session_state.page = "📊 Dashboard Utama"

    selected_page = st.sidebar.radio(
        "Pilih Halaman",
        options=["📊 Dashboard Utama", "📈 Dashboard Internal", "🔐 Admin"],
        key="page"
    )

    st.sidebar.markdown("---")
    st.sidebar.info("""
    **Dashboard IKPA**  
    Indikator Kinerja Pelaksanaan Anggaran  
    KPPN Baturaja  

    📧 Support: ameer.noor@kemenkeu.go.id
    """)

    # ===============================
    # 🔹 Routing Halaman
    # ===============================
    if st.session_state.page == "📊 Dashboard Utama":
        page_dashboard()

    elif st.session_state.page == "📈 Dashboard Internal":
        page_trend()

    elif st.session_state.page == "🔐 Admin":
        page_admin()

# ===============================
# 🔹 ENTRY POINT
# ===============================
if __name__ == "__main__":
    main()
